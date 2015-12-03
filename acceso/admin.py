from django.contrib import admin
from .models import Ciclo,Componente, Modulo, Actividad, Encargado, Entregables, Valor, Evidencia, Corte, CargaMasiva, CargasMasivas
from django.http import HttpResponse
from conf import settings
import openpyxl
from gestor.models import Gestor
import time
import datetime
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font

t = Style(font=Font(name='Calibri',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='C9C9C9',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

co = Style(font=Font(name='Calibri',size=11),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

v = Style(font=Font(name='Calibri',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='E4F5E1',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

vc = Style(font=Font(name='Calibri',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='D4F5CE',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

class ModuloAdmin(admin.ModelAdmin):
    list_display = ('nombre','descripcion')

class EvidenciaAdmin(admin.ModelAdmin):
    list_display = ('radicado','ciclo','encargado')
    search_fields = ('radicado__numero',)

def reasignacion_actividades(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Carga Masiva.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Reasignacion Actividades"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Reasignacion Actividades'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['RADICADO',30]),
                   tuple(['ID',30]),
                   tuple(['CEDULA',30]),
                   tuple(['NOMBRE',30]),
                   tuple(['RESULTADO',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        archivo_masivo = openpyxl.load_workbook(settings.MEDIA_ROOT+'/'+unicode(archivo_queryset.archivo))

        hoja1_masivo = archivo_masivo.get_sheet_by_name('Hoja1')

        i = 0

        for fila in hoja1_masivo.rows:
            i += 1
            if i > 1:
                proceso =""

                if fila[0].value != None:
                    if Evidencia.objects.filter(radicado__numero=fila[0].value).filter(actividad__id=fila[1].value).count() == 1:
                        evidencia = Evidencia.objects.filter(actividad__id=fila[1].value).get(radicado__numero=fila[0].value)
                        if fila[2].value != None:
                            try:
                                cedula = long(fila[2].value)
                            except ValueError:
                                proceso = "El numero de cedula solo puede contener numeros"
                            else:
                                if Gestor.objects.filter(cedula=cedula).count() == 1:
                                    gestor = Gestor.objects.get(cedula=cedula)
                                    evidencia.gestor = gestor
                                    evidencia.save()
                                    proceso = "Cambiado correctamente"
                                elif Gestor.objects.filter(cedula=cedula).count() == 0:
                                    proceso = "No existe el gestor"
                                else:
                                    proceso = "Existe mas de un gestor con el mismo numero de cedula"
                        else:
                            proceso = "Numero de cedula vacia"
                    elif Evidencia.objects.filter(radicado__numero=fila[0].value).filter(actividad__id=fila[1].value).count() == 0:
                        proceso = "No existe el radicado"
                    else:
                        proceso = "Hay mas de un radicado con el mismo numero"
                else:
                    proceso = "El campo de radicado esta vacio"

                row_num += 1
                row = [
                    fila[0].value,
                    fila[1].value,
                    fila[2].value,
                    fila[3].value,
                    proceso
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
reasignacion_actividades.short_description = "Reasignacion de actividades"

def reasignacion_actividades_corte(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Carga Masiva.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Reasignacion Corte"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Reasignacion Corte'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['RADICADO',30]),
                   tuple(['ID',30]),
                   tuple(['CORTE',30]),
                   tuple(['RESULTADO',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        archivo_masivo = openpyxl.load_workbook(settings.MEDIA_ROOT+'/'+unicode(archivo_queryset.archivo))

        hoja1_masivo = archivo_masivo.get_sheet_by_name('Hoja1')

        i = 0

        for fila in hoja1_masivo.rows:
            i += 1
            if i > 1:
                proceso =""

                if fila[0].value != None:
                    if Evidencia.objects.filter(radicado__numero=fila[0].value).filter(actividad__id=fila[1].value).count() == 1:
                        evidencia = Evidencia.objects.filter(radicado__numero=fila[0].value)
                        if fila[1].value != None:
                            if evidencia.filter(actividad__id=fila[1].value).count() == 1:
                                evidencia = evidencia.filter(actividad__id=fila[1].value)[0]
                                if fila[1].value != None:
                                    corte = Corte.objects.filter(id=fila[2].value)
                                    if corte.count() == 1:
                                        corte = corte[0]
                                        evidencia.corte = corte
                                        evidencia.save()
                                    else:
                                        proceso = "Campo de ID corte invalido"
                                else:
                                    proceso = "Campo de ID corte vacio"
                            else:
                                proceso = "Campo de ID invalido"
                        else:
                            proceso = "Campo de ID actividad vacio"
                    elif Evidencia.objects.filter(radicado__numero=fila[0].value).filter(actividad__id=fila[1].value).count() == 0:
                        proceso = "No existe el radicado"
                    else:
                        proceso = "Hay mas de un radicado con el mismo numero"
                else:
                    proceso = "El campo de radicado esta vacio"

                row_num += 1
                row = [
                    fila[0].value,
                    fila[1].value,
                    fila[2].value,
                    proceso
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
reasignacion_actividades_corte.short_description = "Reasignacion de actividades corte"

class CargasMasivasAdmin(admin.ModelAdmin):
    list_display = ['id','archivo']
    ordering = ['archivo']
    actions = [reasignacion_actividades,reasignacion_actividades_corte]

admin.site.register(Ciclo)
admin.site.register(Componente)
admin.site.register(Modulo,ModuloAdmin)
admin.site.register(Actividad)
admin.site.register(Encargado)
admin.site.register(Entregables)
admin.site.register(Valor)
admin.site.register(Evidencia,EvidenciaAdmin)
admin.site.register(Corte)
admin.site.register(CargaMasiva)
admin.site.register(CargasMasivas,CargasMasivasAdmin)