import openpyxl
import os

path = os.getcwd() + '\\'
archivo = openpyxl.load_workbook(path+'Libro RED.xlsx')
hoja = archivo.get_sheet_by_name('RED FORMACION')
columnas = [{'columna':'N','path':'Nivel 1/Sesion 1/'},
            {'columna':'O','path':'Nivel 1/Sesion 2/'},
            {'columna':'P','path':'Nivel 1/Sesion 3/'},
            {'columna':'Q','path':'Nivel 1/Sesion 4/'},
            {'columna':'AA','path':'Nivel 2/Sesion 1/'},
            {'columna':'AB','path':'Nivel 2/Sesion 1/'},
            {'columna':'AH','path':'Nivel 3/Sesion 1/'},
            {'columna':'AI','path':'Nivel 3/Sesion 1/'},
            {'columna':'AJ','path':'Nivel 3/Sesion 1/'},
            {'columna':'AR','path':'Nivel 4/Sesion 1/'},
            {'columna':'AS','path':'Nivel 4/Sesion 1/'},
            {'columna':'AT','path':'Nivel 4/Sesion 1/'},
            {'columna':'AU','path':'Nivel 4/Sesion 1/'},
            {'columna':'AV','path':'Nivel 4/Sesion 1/'},]
i=5
for fila in hoja.rows:
    for columna in columnas:
        if hoja.cell(columna['columna']+str(i)).value != None:
            y = path + columna['path'] + hoja.cell(columna['columna']+str(i)).value
            hoja.cell(columna['columna']+str(i)).value = "X"
            hoja.cell(columna['columna']+str(i)).hyperlink = "file:////"+y

    i += 1
archivo.save(path+'Libro RED OK.xlsx')
input("Ya puedes cerrar la ventana")