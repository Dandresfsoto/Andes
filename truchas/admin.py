#!/usr/bin/env python
# -*- coding: utf-8 -*-
from django.contrib import admin
from .models import ParticipanteEscuelaTicTrucho, CargasMasivas
from formador.models import Formador
from municipio.models import Municipio
from django.http import HttpResponse
from conf import settings
import openpyxl
from formacion.models import ParticipanteEscuelaTic, Grupo
import time
import datetime
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font
from .models import CodigoMasivo
from region.models import Region
from formador.models import Formador
import pythoncom
from win32com import client
from win32com.client import DispatchEx
from win32com import *
from django.shortcuts import HttpResponseRedirect
from .models import Nivel1_Sesion1_1,Nivel1_Sesion1_2,Nivel1_Sesion1_3,Nivel1_Sesion1_4,Nivel1_Sesion1_5,Nivel1_Sesion1_6,Nivel1_Sesion1_7,Nivel1_Sesion1_8,Nivel1_Sesion1_9,Nivel1_Sesion1_10,Nivel1_Sesion1_11,Nivel1_Sesion1_12,Nivel1_Sesion1_REDA
from fdfgen import forge_fdf
import os
import random
from formacion.models import SoporteEntregableDocente, ParticipanteDocente, EvidenciaDocentes, EntregableDocentes, ValorDocente, EvidenciaEscuelaTic
from django.core.files import File
from truchas.models import Nivel1_Sesion2,Nivel1_Sesion3,Nivel1_Sesion4_preguntas_aleatorias,Nivel1_Sesion4_1,Nivel1_Sesion4_2,Nivel1_Sesion4_3,Nivel1_Sesion4_4,Nivel1_Sesion4_5
from formacion.models import EvidenciaDocentes
from truchas.models import Nivel4_Sesion1_1, Nivel4_Sesion1_2,Nivel4_Sesion2_1, Nivel4_Sesion2_2 ,Nivel4_Sesion2_3
from truchas.models import Nivel4_Sesion3_1, Nivel4_Sesion3_2, Nivel4_Sesion3_3, Nivel4_Sesion3_4, Nivel4_Sesion3_5, Nivel4_Sesion3_6
from truchas.models import ParticipanteN1_S2, CodigoMasivoN1_S2
#from pptx import Presentation
from truchas.models import Nivel1_Sesion2_1,Nivel1_Sesion2_2,Nivel1_Sesion2_3,Nivel1_Sesion2_4
from truchas.models import Nivel3_Sesion3_1, Nivel3_Sesion3_2
from truchas.models import Nivel3_Sesion1_1, Nivel3_Sesion1_2, Nivel3_Sesion2_1
from truchas.models import CodigoMasivo_Docentes, CodigoMasivo_Proyectoss
from truchas.models import ParticipanteDocenteMasivo, ParticipanteProyectoMasivos
from truchas.models import CargaMasiva_n3_n4
#from docx import Document
from zipfile import ZipFile
from formacion.models import GrupoDocentes
from PyPDF2 import PdfFileWriter, PdfFileReader
import shutil

t = Style(font=Font(name='Calibri',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='C9C9C9',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

co = Style(font=Font(name='Calibri',size=11),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

v = Style(font=Font(name='Calibri',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='E4F5E1',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

vc = Style(font=Font(name='Calibri',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='D4F5CE',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

def validateEmail( email ):
    from django.core.validators import validate_email
    from django.core.exceptions import ValidationError
    try:
        validate_email( email )
        return True
    except ValidationError:
        return False

def _removeNonAscii(s): return "".join(i for i in s if ord(i)<128)

def carga_participantes(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Carga Masiva.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Carga Masiva Participantes"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Carga Masiva Participantes'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['CODIGO MASIVO',30]),
                   tuple(['REGION',30]),
                   tuple(['DEPARTAMENTO',30]),
                   tuple(['MUNICIPIO',30]),
                   tuple(['INSTITUCION',30]),
                   tuple(['GRUPO',30]),
                   tuple(['GRUPO',30]),
                   tuple(['NOMBRE FORMADOR',30]),
                   tuple(['CEDULA FORMADOR',30]),
                   tuple(['NOMBRES',30]),
                   tuple(['APELLIDOS',30]),
                   tuple(['CEDULA',60]),
                   tuple(['GENERO',30]),
                   tuple(['NIVEL EDUCATIVO',60]),
                   tuple(['CELULAR',30]),
                   tuple(['CORREO',30]),
                   tuple(['TIPO POBLACION',30]),
                   tuple(['CODIGO ANSPE',30]),
                   tuple(['TIPO PROYECTO',30]),
                   tuple(['GRUPO DE CONFORMACION',30]),
                   tuple(['RESULTADO',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        archivo_masivo = openpyxl.load_workbook(settings.MEDIA_ROOT+'/'+unicode(archivo_queryset.archivo))

        hoja1_masivo = archivo_masivo.get_sheet_by_name('MATRIZ PADRES DE FAMILIA-INTERV')

        i = 0

        for fila in hoja1_masivo.rows:
            i += 1
            if i > 3:
                proceso =""
                try:
                    cedula = fila[10].value.replace('.','').replace(',','')
                except:
                    cedula = fila[10].value

                nombres = fila[8].value
                apellidos = fila[9].value

                if CodigoMasivo.objects.filter(codigo=int(fila[0].value)).count() == 0:
                    x = CodigoMasivo(codigo=int(fila[0].value))
                    x.save()

                codigo_masivo = CodigoMasivo.objects.get(codigo=int(fila[0].value))
                cantidad = ParticipanteEscuelaTicTrucho.objects.filter(codigo_masivo__codigo=int(fila[0].value)).count()

                if codigo_masivo.generado == False and cantidad <= 20:
                    if cedula == "" or cedula == None:
                        proceso = "El campo de cedula esta vacio"
                    else:
                        try:
                            long(cedula)
                        except ValueError:
                            proceso = "El numero de cedula es invalido"
                        else:
                            if 0 != 0:
                                proceso = "La Cedula ya se encuentra registrada (Modulo Principal)"
                            else:
                                if ParticipanteEscuelaTicTrucho.objects.filter(cedula=cedula).count() != 0:
                                    proceso = "La Cedula ya se encuentra registrada (Modulo Actual)"
                                else:
                                    if nombres == "" or nombres == None:
                                        proceso = "El campo de nombres esta vacio"
                                    else:
                                        if apellidos == "" or apellidos == None:
                                            proceso = "El campo de apellidos esta vacio"
                                        else:
                                            if Formador.objects.filter(cedula=fila[7].value).count() == 0:
                                                proceso = "No existe el formador"
                                            else:
                                                formador = Formador.objects.filter(cedula=fila[7].value)[0]
                                                if Grupo.objects.filter(formador__id=formador.id).count() == 0:
                                                    proceso = "El formador no tiene grupos asignados"
                                                else:
                                                    grupo = Grupo.objects.filter(formador__id=formador.id)[0]
                                                    proceso = "Registrado Correctamente"
                                                    nuevo = ParticipanteEscuelaTicTrucho()
                                                    nuevo.codigo_masivo = codigo_masivo

                                                    if fila[1].value == 1:
                                                        nuevo.region = Region.objects.get(id=1)
                                                    if fila[1].value == 4:
                                                        nuevo.region = Region.objects.get(id=2)

                                                    nuevo.departamento = fila[2].value
                                                    nuevo.municipio = fila[3].value
                                                    nuevo.institucion = fila[4].value
                                                    nuevo.grupo = grupo
                                                    nuevo.formador = grupo.formador

                                                    nuevo.nombres = fila[8].value
                                                    nuevo.apellidos = fila[9].value
                                                    nuevo.cedula = long(cedula)
                                                    if fila[11].value == None or fila[11].value == "":
                                                        nuevo.genero = "Masculino"
                                                        proceso = "Registrado Correctamente - Genero Masculino por defecto"
                                                    else:
                                                        nuevo.genero = fila[11].value
                                                    nuevo.nivel_educativo = fila[12].value
                                                    nuevo.telefono = fila[13].value
                                                    if validateEmail(fila[14].value):
                                                        nuevo.correo = fila[14].value
                                                    nuevo.poblacion = fila[15].value
                                                    nuevo.codigo_anspe = fila[16].value
                                                    nuevo.tipo_proyecto = fila[17].value
                                                    nuevo.grupo_conformacion = fila[18].value
                                                    nuevo.save()

                elif codigo_masivo.generado == True:
                    proceso = "El listado ya fue generado y no se pueden agregar mas participantes"
                else:
                    proceso = "Hay mas de 20 participantes en el listado"

                row_num += 1
                row = [
                    fila[0].value,
                    fila[1].value,
                    fila[2].value,
                    fila[3].value,
                    fila[4].value,
                    fila[5].value,
                    fila[6].value,
                    fila[7].value,
                    fila[8].value,
                    fila[9].value,
                    fila[10].value,
                    fila[11].value,
                    fila[12].value,
                    fila[13].value,
                    fila[14].value,
                    fila[15].value,
                    fila[16].value,
                    fila[17].value,
                    fila[18].value,
                    proceso
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
carga_participantes.short_description = "Cargar participantes"

def carga_n1_s2(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Carga Masiva Grupos N1 S2.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Carga Masiva Grupos"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Carga Masiva Grupos'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['CEDULA',30]),tuple(['CODIGO',30]),tuple(['RESULTADO',30])]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        archivo_masivo = openpyxl.load_workbook(settings.MEDIA_ROOT+'/'+unicode(archivo_queryset.archivo))

        hoja1_masivo = archivo_masivo.get_sheet_by_name('Hoja1')

        i = 0

        for fila in hoja1_masivo.rows:
            i += 1
            if i > 1:
                proceso =""

                if CodigoMasivoN1_S2.objects.filter(codigo = fila[1].value).count() == 0:
                    codigo = CodigoMasivoN1_S2()
                    codigo.codigo = fila[1].value
                    codigo.save()
                else:
                    codigo = CodigoMasivoN1_S2.objects.get(codigo = fila[1].value)

                try:
                    docente = ParticipanteDocente.objects.get(cedula = fila[0].value)
                except:
                    proceso = "No existe el docente con el documento de identidad"
                else:
                    nuevo = ParticipanteN1_S2(codigo_masivo=codigo,participante=docente)
                    nuevo.save()
                    proceso = "Cargado"

                row_num += 1
                row = [
                    fila[0].value,
                    fila[1].value,
                    proceso
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
carga_n1_s2.short_description = "Cargar grupos N1 - S2"


def eliminar_actividades(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Actividades Eliminadas.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Eliminar Actividades"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Eliminar Actividades'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['CEDULA',30]),
                   tuple(['ID',30]),
                   tuple(['RESULTADO',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        archivo_masivo = openpyxl.load_workbook(settings.MEDIA_ROOT+'/'+unicode(archivo_queryset.archivo))

        hoja1_masivo = archivo_masivo.get_sheet_by_name('Hoja1')

        i = 0

        for fila in hoja1_masivo.rows:
            i += 1
            if i > 1:
                proceso =""

                try:
                    participante = EvidenciaDocentes.objects.filter(participante__cedula=fila[0].value).get(entregable__id=fila[1].value)
                except:
                    proceso = "Error"
                else:
                    proceso = "Soporte Eliminado"
                    try:
                        soporte = SoporteEntregableDocente.objects.get(id=participante.soporte.id)
                    except:
                        proceso = "No existe la actividad"
                    else:
                        soporte.soporte = None
                        soporte.save()
                        participante.soporte = None
                        participante.save()

                row_num += 1
                row = [
                    fila[0].value,
                    fila[1].value,
                    proceso
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
eliminar_actividades.short_description = "Eliminar Actividades"

def verificar_listados(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Actividades Eliminadas.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Verificar Listados"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Verificar Listados'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['CEDULA',30]),
                   tuple(['RESULTADO',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        archivo_masivo = openpyxl.load_workbook(settings.MEDIA_ROOT+'/'+unicode(archivo_queryset.archivo))

        hoja1_masivo = archivo_masivo.get_sheet_by_name('Hoja1')

        i = 0

        for fila in hoja1_masivo.rows:
            i += 1
            if i > 1:
                proceso1 =""
                proceso2 =""

                try:
                    participante = EvidenciaEscuelaTic.objects.filter(participante__cedula=fila[0].value)
                except:
                    proceso1 = "Numero de cedula no encontrado"
                else:
                    try:
                        soporte = participante.get(entregable__id=5).soporte.soporte
                    except:
                        proceso1 = "No"
                    else:
                        if unicode(soporte) == '':
                            proceso1 = "No"
                        else:
                            proceso1 = "Si"

                    try:
                        soporte = participante.get(entregable__id=9).soporte.soporte
                    except:
                        proceso1 = "No"
                    else:
                        if unicode(soporte) == '':
                            proceso2 = "No"
                        else:
                            proceso2 = "Si"

                row_num += 1
                row = [
                    fila[0].value,
                    proceso1,
                    proceso2
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
verificar_listados.short_description = "Verificar Listados Padres"

def reporte_formadores(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Reporte Formadores.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Reporte Formadores"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Reporte Formadores'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['Formador',30]),
                   tuple(['Región',30]),
                   tuple(['Nivel 1 - Sesión 1',30]),
                   tuple(['Nivel 1 - Sesión 2',30]),
                   tuple(['Nivel 1 - Sesión 3',30]),
                   tuple(['Nivel 1 - Sesión 4',30]),
                   tuple(['Nivel 2 - Sesión 1',30]),
                   tuple(['Nivel 2 - Sesión 2',30]),
                   tuple(['Nivel 3 - Sesión 1',30]),
                   tuple(['Nivel 3 - Sesión 2',30]),
                   tuple(['Nivel 3 - Sesión 3',30]),
                   tuple(['Nivel 4 - Sesión 1',30]),
                   tuple(['Nivel 4 - Sesión 2',30]),
                   tuple(['Nivel 4 - Sesión 3',30]),
                   tuple(['Nivel 4 - Sesión 4',30]),
                   tuple(['Nivel 4 - Sesión 5',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        validos = [1,3,5,7,21,23,29,31,33,41,43,45,47,49]

        for formador in Formador.objects.filter(tipo=1):
                row_num += 1
                row = [
                    formador.nombre,
                    formador.region.nombre,
                    EvidenciaDocentes.objects.filter(participante__formador__id=formador.id).filter(entregable__id=1).exclude(soporte=None).count(),
                    EvidenciaDocentes.objects.filter(participante__formador__id=formador.id).filter(entregable__id=3).exclude(soporte=None).count(),
                    EvidenciaDocentes.objects.filter(participante__formador__id=formador.id).filter(entregable__id=5).exclude(soporte=None).count(),
                    EvidenciaDocentes.objects.filter(participante__formador__id=formador.id).filter(entregable__id=7).exclude(soporte=None).count(),
                    EvidenciaDocentes.objects.filter(participante__formador__id=formador.id).filter(entregable__id=21).exclude(soporte=None).count(),
                    EvidenciaDocentes.objects.filter(participante__formador__id=formador.id).filter(entregable__id=23).exclude(soporte=None).count(),
                    EvidenciaDocentes.objects.filter(participante__formador__id=formador.id).filter(entregable__id=29).exclude(soporte=None).count(),
                    EvidenciaDocentes.objects.filter(participante__formador__id=formador.id).filter(entregable__id=31).exclude(soporte=None).count(),
                    EvidenciaDocentes.objects.filter(participante__formador__id=formador.id).filter(entregable__id=33).exclude(soporte=None).count(),
                    EvidenciaDocentes.objects.filter(participante__formador__id=formador.id).filter(entregable__id=41).exclude(soporte=None).count(),
                    EvidenciaDocentes.objects.filter(participante__formador__id=formador.id).filter(entregable__id=43).exclude(soporte=None).count(),
                    EvidenciaDocentes.objects.filter(participante__formador__id=formador.id).filter(entregable__id=45).exclude(soporte=None).count(),
                    EvidenciaDocentes.objects.filter(participante__formador__id=formador.id).filter(entregable__id=47).exclude(soporte=None).count(),
                    EvidenciaDocentes.objects.filter(participante__formador__id=formador.id).filter(entregable__id=49).exclude(soporte=None).count(),
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
reporte_formadores.short_description = "Reporte de Formadores Tipo 1"

def reporte_docentes(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Reporte Docentes.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Reporte Docentes"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Reporte Docentes'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['Region',30]),

                   tuple(['Municipio',30]),
                   tuple(['Departamento',30]),
                   tuple(['Formador',30]),
                   tuple(['Cedula Formador',30]),

                   tuple(['Nombres',30]),
                   tuple(['Apellidos',30]),
                   tuple(['Cedula',30]),
                   tuple(['Nivel 1 - Sesión 1',30]),
                   tuple(['Nivel 1 - Sesión 2',30]),
                   tuple(['Nivel 1 - Sesión 3',30]),
                   tuple(['Nivel 1 - Sesión 4',30]),
                   tuple(['Nivel 2 - Sesión 1',30]),
                   tuple(['Nivel 2 - Sesión 2',30]),
                   tuple(['Nivel 3 - Sesión 1',30]),
                   tuple(['Nivel 3 - Sesión 2',30]),
                   tuple(['Nivel 3 - Sesión 3',30]),
                   tuple(['Nivel 4 - Sesión 1',30]),
                   tuple(['Nivel 4 - Sesión 2',30]),
                   tuple(['Nivel 4 - Sesión 3',30]),
                   tuple(['Nivel 4 - Sesión 4',30]),
                   tuple(['Nivel 4 - Sesión 5',30]),

                   tuple(['Nivel 1 - Cuestionario Explorando portales educativos',30]),
                   tuple(['Nivel 1 - Participación  en el foro socializando la secuencia didáctica con reda (presentación en Power Point)',30]),
                   tuple(['Nivel 1 - Documento resultado de la aplicación de la prueba',30]),
                   tuple(['Nivel 1 - Guía 4: diseñando el cuestionario diagnóstico',30]),
                   tuple(['Nivel 1 - Aprobación al menos con un 60% de la evaluación del Nivel 1',30]),
                   tuple(['Nivel 2 - Gestor de proyectos',30]),
                   tuple(['Nivel 2 - Documento refuerzo',30]),
                   tuple(['Nivel 3 - Presentación en Prezi,  Powtoon, Slideshare o Power Point  con los resultados de la ejecución del Proyecto Educativo TIC.',30]),

                   tuple(['Nivel 3 - 1. Un documento en Word  con las tres conclusiones. 2. Un documento de Excel con los resultados de la prueba estandarizada.',30]),
                   tuple(['Nivel 3 - Cuestionario: Evaluando competencias adquiridas.',30]),
                   tuple(['Nivel 4 - Gestor de proyectos',30]),

                   #tuple(['Nivel 3 - v 12',30]),
                   #tuple(['Nivel 3 - v 13',30]),
                   #tuple(['Nivel 3 - v 14',30]),
                   #tuple(['Nivel 3 - v 15',30]),
                   #tuple(['Nivel 3 - v 16',30]),
                   #tuple(['Nivel 3 - v 17',30]),

                   #tuple(['Nivel 4 - v 18',30]),
                   #tuple(['Nivel 4 - v 19',30]),
                   #tuple(['Nivel 4 - v 20',30]),
                   #tuple(['Nivel 4 - v 21',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]



        for participante in ParticipanteDocente.objects.all():
                evidencias = EvidenciaDocentes.objects.filter(participante__id=participante.id)

                try:
                    ac1 = evidencias.get(entregable__id=1).soporte.soporte
                except:
                    ac1 = ""

                try:
                    ac2 = evidencias.get(entregable__id=3).soporte.soporte
                except:
                    ac2 = ""

                try:
                    ac3 = evidencias.get(entregable__id=5).soporte.soporte
                except:
                    ac3 = ""

                try:
                    ac4 = evidencias.get(entregable__id=7).soporte.soporte
                except:
                    ac4 = ""

                try:
                    ac5 = evidencias.get(entregable__id=21).soporte.soporte
                except:
                    ac5 = ""

                try:
                    ac6 = evidencias.get(entregable__id=23).soporte.soporte
                except:
                    ac6 = ""

                try:
                    ac7 = evidencias.get(entregable__id=29).soporte.soporte
                except:
                    ac7 = ""

                try:
                    ac8 = evidencias.get(entregable__id=31).soporte.soporte
                except:
                    ac8 = ""

                try:
                    ac9 = evidencias.get(entregable__id=33).soporte.soporte
                except:
                    ac9 = ""


                try:
                    ac10 = evidencias.get(entregable__id=41).soporte.soporte
                except:
                    ac10 = ""

                try:
                    ac11 = evidencias.get(entregable__id=43).soporte.soporte
                except:
                    ac11 = ""

                try:
                    ac12 = evidencias.get(entregable__id=45).soporte.soporte
                except:
                    ac12 = ""

                try:
                    ac13 = evidencias.get(entregable__id=47).soporte.soporte
                except:
                    ac13 = ""

                try:
                    ac14 = evidencias.get(entregable__id=49).soporte.soporte
                except:
                    ac14 = ""



                try:
                    av1 = evidencias.get(entregable__id=11).soporte.soporte
                except:
                    av1 = ""


                try:
                    av2 = evidencias.get(entregable__id=13).soporte.soporte
                except:
                    av2 = ""

                try:
                    av3 = evidencias.get(entregable__id=15).soporte.soporte
                except:
                    av3 = ""


                try:
                    av4 = evidencias.get(entregable__id=17).soporte.soporte
                except:
                    av4 = ""


                try:
                    av5 = evidencias.get(entregable__id=9).soporte.soporte
                except:
                    av5 = ""


                try:
                    av6 = evidencias.get(entregable__id=60).soporte.soporte
                except:
                    av6 = ""


                try:
                    av7 = evidencias.get(entregable__id=27).soporte.soporte
                except:
                    av7 = ""


                try:
                    av8 = evidencias.get(entregable__id=35).soporte.soporte
                except:
                    av8 = ""



                try:
                    av9 = evidencias.get(entregable__id=37).soporte.soporte
                except:
                    av9 = ""



                try:
                    av10 = evidencias.get(entregable__id=39).soporte.soporte
                except:
                    av10 = ""


                try:
                    av11 = evidencias.get(entregable__id=60).soporte.soporte
                except:
                    av11 = ""







                row_num += 1
                row = [
                    participante.formador.region.nombre,

                    participante.grupo.municipio.nombre,
                    participante.grupo.municipio.departamento.nombre,
                    participante.formador.nombre,
                    participante.formador.cedula,

                    participante.nombres,
                    participante.apellidos,
                    participante.cedula,


                    unicode(ac1),
                    unicode(ac2),
                    unicode(ac3),
                    unicode(ac4),
                    unicode(ac5),
                    unicode(ac6),
                    unicode(ac7),
                    unicode(ac8),
                    unicode(ac9),
                    unicode(ac10),
                    unicode(ac11),
                    unicode(ac12),
                    unicode(ac13),
                    unicode(ac14),

                    unicode(av1),
                    unicode(av2),
                    unicode(av3),
                    unicode(av4),
                    unicode(av5),
                    unicode(av6),
                    unicode(av7),
                    unicode(av8),
                    unicode(av9),
                    unicode(av10),
                    unicode(av11),



                    #evidencias.get(entregable__id=11).soporte,
                    #evidencias.get(entregable__id=11).soporte,
                    #evidencias.get(entregable__id=13).soporte,
                    #evidencias.get(entregable__id=15).soporte,
                    #evidencias.get(entregable__id=17).soporte,
                    #evidencias.get(entregable__id=17).soporte,
                    #evidencias.get(entregable__id=9).soporte,
                    #evidencias.get(entregable__id=9).soporte,
                    #evidencias.get(entregable__id=25).soporte,
                    #evidencias.get(entregable__id=26).soporte,
                    #evidencias.get(entregable__id=27).soporte,
                    #evidencias.get(entregable__id=35).soporte,
                    #evidencias.get(entregable__id=36).soporte,
                    #evidencias.get(entregable__id=37).soporte,
                    #evidencias.get(entregable__id=38).soporte,
                    #evidencias.get(entregable__id=39).soporte,
                    #evidencias.get(entregable__id=40).soporte,
                    #evidencias.get(entregable__id=51).soporte,
                    #evidencias.get(entregable__id=53).soporte,
                    #evidencias.get(entregable__id=55).soporte,
                    #evidencias.get(entregable__id=59).soporte,

                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
reporte_docentes.short_description = "Reporte de Docentes Vinculados"

def reporte_formadores_tipo2(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Reporte Formadores.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Reporte Formadores"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Reporte Formadores'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['Formador',30]),
                   tuple(['Región',30]),
                   tuple(['Padres - Sesión 1',30]),
                   tuple(['Padres - Sesión 2',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        validos = [5,9]

        for formador in Formador.objects.filter(tipo=2):
                row_num += 1
                row = [
                    formador.nombre,
                    formador.region.nombre,
                    EvidenciaEscuelaTic.objects.filter(participante__formador__id=formador.id).filter(entregable__id=5).exclude(soporte=None).count(),
                    EvidenciaEscuelaTic.objects.filter(participante__formador__id=formador.id).filter(entregable__id=9).exclude(soporte=None).count(),
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
reporte_formadores_tipo2.short_description = "Reporte de Formadores Tipo 2"

def estructura_cedulas(modeladmin,request,queryset):
    participantes = ParticipanteDocente.objects.all()

    listados = [{'sesion':1,'path':'Nivel 1/Sesion 1'},{'sesion':3,'path':'Nivel 1/Sesion 2'},
                {'sesion':5,'path':'Nivel 1/Sesion 3'},{'sesion':7,'path':'Nivel 1/Sesion 4'},
                {'sesion':21,'path':'Nivel 2/Sesion 1'},{'sesion':23,'path':'Nivel 2/Sesion 2'}]

    for participante in participantes:
        if not os.path.exists('C:/Temp/Estructura/'+unicode(participante.cedula)):
            os.mkdir('C:/Temp/Estructura/'+unicode(participante.cedula))
        evidencias = EvidenciaDocentes.objects.filter(participante__cedula=participante.cedula)
        for listado in listados:
            try:
                path = evidencias.get(entregable__id=listado['sesion']).soporte.soporte.path
            except:
                pass
            else:
                os.makedirs('C:/Temp/Estructura/'+unicode(participante.cedula)+'/'+listado['path'])
                shutil.copy2(path,'C:/Temp/Estructura/'+unicode(participante.cedula)+'/'+listado['path'])
estructura_cedulas.short_description = "Exportar Estructura Cedulas"


class CargasMasivasAdmin(admin.ModelAdmin):
    list_display = ['id','archivo']
    ordering = ['archivo']
    actions = [carga_participantes,eliminar_actividades,verificar_listados,carga_n1_s2,estructura_cedulas]
admin.site.register(CargasMasivas, CargasMasivasAdmin)
admin.site.register(ParticipanteEscuelaTicTrucho)
admin.site.register(ParticipanteN1_S2)


def listado(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Reporte Formadores.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Listado"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Listado'

        celda = hoja1.cell('E3')
        celda.value = 'Listado'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['Grupo',30]),
                   tuple(['Cedulas',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]



        for codigo in CodigoMasivoN1_S2.objects.all():
                row_num += 1
                row = [
                    codigo.codigo,
                    unicode(ParticipanteN1_S2.objects.filter(codigo_masivo = codigo).values_list('participante__cedula',flat=True))
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
listado.short_description = "Listado en filas"




'''
def generar_sesion2_n1(modeladmin,request,queryset):
    for codigo in queryset:
        participantes = ParticipanteN1_S2.objects.filter(codigo_masivo=codigo)
        aleatorio = str(random.randint(1,78))
        presentacion = Presentation('C:\\Temp\PRESENTACIONES\\'+aleatorio+'.pptx')
        propiedades = presentacion.core_properties
        propiedades.author = _removeNonAscii(participantes[0].participante.nombres)
        propiedades.created = datetime.datetime(2015,random.randint(9,12),random.randint(1,30),random.randint(0,23),random.randint(0,59),random.randint(0,59))
        propiedades.last_modified_by = _removeNonAscii(participantes[0].participante.nombres)
        propiedades.modified = datetime.datetime(2015,random.randint(9,12),random.randint(1,30),random.randint(0,23),random.randint(0,59),random.randint(0,59))

        presentacion.slides[0].shapes.title.text = "Socializando la Secuencia Didactica"
        nombres = ""
        for participante in participantes:
            nombres += participante.participante.nombres+" "+participante.participante.apellidos+" - "

        presentacion.slides[0].shapes.placeholders[1].text = nombres[:len(nombres)-3]

        slide_1 = presentacion.slides.add_slide(presentacion.slide_layouts[1])
        slide_1.shapes.placeholders[0].text = "Resultados de lo Planeado y Ejecutado"
        slide_1.shapes.placeholders[1].text = Nivel1_Sesion2_1.objects.all().order_by('?').first().respuesta

        slide_2 = presentacion.slides.add_slide(presentacion.slide_layouts[1])
        slide_2.shapes.placeholders[0].text = "Lo Planeado y no Ejecutado"
        slide_2.shapes.placeholders[1].text = Nivel1_Sesion2_2.objects.all().order_by('?').first().respuesta

        slide_3 = presentacion.slides.add_slide(presentacion.slide_layouts[1])
        slide_3.shapes.placeholders[0].text = "Lo no Planeado y Ejecutado"
        slide_3.shapes.placeholders[1].text = Nivel1_Sesion2_3.objects.all().order_by('?').first().respuesta

        opciones = ['En la actividad participaron ','Participaron ','Se seleccionaron ','La actividad fue conformada por ',
                    'En la secuencia didactica participaron ']

        slide_4 = presentacion.slides.add_slide(presentacion.slide_layouts[1])
        slide_4.shapes.placeholders[0].text = "Numero de estudiantes participantes"
        slide_4.shapes.placeholders[1].text = opciones[random.randint(0,len(opciones)-1)]+str(random.randint(10,40))+" estudiantes."

        slide_5 = presentacion.slides.add_slide(presentacion.slide_layouts[1])
        slide_5.shapes.placeholders[0].text = "Resultados Evaluativos"
        slide_5.shapes.placeholders[1].text = Nivel1_Sesion2_4.objects.all().order_by('?').first().respuesta

        presentacion.save('C:\\Temp\\Descarga\\'+str(participantes[0].participante.cedula)+'.pptx')
generar_sesion2_n1.short_description = "Generar presentaciones"
'''

class CodigoMasivoN1_S2Admin(admin.ModelAdmin):
    list_display = ['codigo']
    ordering = ['codigo']
    actions = [listado]
admin.site.register(CodigoMasivoN1_S2,CodigoMasivoN1_S2Admin)


def generar_listas_docentes(modeladmin,request,queryset):
    pythoncom.CoInitializeEx(pythoncom.COINIT_MULTITHREADED)
    xl = win32com.client.dynamic.Dispatch('Excel.Application')
    xl.DisplayAlerts = True
    xl.Visible = 0

    lista = xl.Workbooks.Open(settings.STATICFILES_DIRS[0]+'/formatos/Listado Asistencia.xlsx')

    i = 2

    for codigo_masivo in queryset:

        parti = ParticipanteDocenteMasivo.objects.filter(codigo_masivo=codigo_masivo)

        sesiones = [{'nivel':2,'sesion':1,'nombre_sesion':'Estructuración del Proyecto','actividades':[('Fundamentando conceptos',4)]},
                        {'nivel':2,'sesion':2,'nombre_sesion':'Cierre de Nivel','actividades':[('Valorando la secuencia didáctica',3),('Evaluando Competencias',1)]},
                        {'nivel':3,'sesion':1,'nombre_sesion':'Preparando la Ejecución','actividades':[('Generando actividades constructivas y participativas',4)]},
                        {'nivel':3,'sesion':2,'nombre_sesion':'Ejecución del Proyecto','actividades':[('Reflexionando para ajustar y mejorar actividad',4)]},
                        {'nivel':3,'sesion':3,'nombre_sesion':'Cierre de Nivel','actividades':[('Como realizar Rubricas',1),('Evaluación formativa',1),('Retroalimentación final',2)]},
                        {'nivel':4,'sesion':1,'nombre_sesion':'Preparando Socialización del proyecto','actividades':[('Preparando presentación del proyecto',4)]},
                        {'nivel':4,'sesion':2,'nombre_sesion':'Uso responsable de las TIC','actividades':[('Compartiendo el proyecto educativo TIC',4)]},
                        {'nivel':4,'sesion':3,'nombre_sesion':'Discutiendo sobe TIC y Medio Ambiente','actividades':[('Gestión de Residuos',4)]},
                        {'nivel':4,'sesion':5,'nombre_sesion':'Comparto experiencia con Comunidad educativa','actividades':[('Comparto con mi Comunidad educativa',4)]}]


        for participantes in map(None,*[iter(parti)]*25):
            for sesion in sesiones:
                fila = 0

                lista.Worksheets("Hoja1").Copy(lista.Worksheets(lista.Worksheets.Count))
                sesion_file = lista.Worksheets('Hoja1 ('+str(i)+')')
                i += 1


                sesion_file.Cells(10,5).Value = sesion['sesion']
                sesion_file.Cells(11,5).Value = sesion['nivel']

                sesion_file.Range("AC8").Value = codigo_masivo.departamento.encode("latin1")
                sesion_file.Range("AC9").Value = codigo_masivo.municipio.encode("latin1")

                sesion_file.Range("AC11").Value = codigo_masivo.formador.encode("latin1")
                sesion_file.Range("AD11").Value = 'C.C. '+codigo_masivo.cedula
                sesion_file.Range("L12").Value = 'Lugar de la Formación (IE): '.encode("latin1")+codigo_masivo.lugar.encode("latin1")

                sesion_file.Range("AF6").Value = sesion['nombre_sesion'].encode("latin1")

                fila_actividad = 7
                for actividad in sesion['actividades']:
                    sesion_file.Range("AF"+str(fila_actividad)).Value = actividad[0].encode("latin1")
                    sesion_file.Range("AH"+str(fila_actividad)).Value = 'Horas:      '+unicode(actividad[1])
                    fila_actividad += 1

                fila = 16
                for participante in participantes:
                    if participante != None:
                        sesion_file.Range("C"+str(fila)).Value = participante.nombre.encode("latin1")
                        sesion_file.Range("J"+str(fila)).Value = participante.cedula
                        sesion_file.Range("AB"+str(fila)).Value = participante.institucion
                        sesion_file.Range("AD"+str(fila)).Value = participante.sede
                        sesion_file.Range("AE"+str(fila)).Value = participante.correo
                        sesion_file.Range("AG"+str(fila)).Value = participante.telefono
                        fila += 1


    lista.Worksheets('Hoja1').Delete()
    lista.ExportAsFixedFormat(0,settings.MEDIA_ROOT+'/Listados Docentes/Docentes.pdf')
    lista.Close(SaveChanges=False)

    xl.Quit()
    pythoncom.CoUninitialize()


    return HttpResponseRedirect('/media/Listados Docentes/Docentes.pdf')
generar_listas_docentes.short_description = "Generar Listas Docentes"



class CodigoMasivo_DocentesAdmin(admin.ModelAdmin):
    list_display = ['codigo','departamento','municipio','formador','subgrupos']
    ordering = ['codigo']
    list_filter = ['region','formador']
    actions = [generar_listas_docentes]

    def subgrupos(self,obj):
        return ParticipanteDocenteMasivo.objects.filter(codigo_masivo=obj.id).values_list('subgrupo').distinct().count()

admin.site.register(CodigoMasivo_Docentes,CodigoMasivo_DocentesAdmin)

class ParticipanteDocenteMasivoAdmin(admin.ModelAdmin):
    list_display = ['cedula','subgrupo']
    search_fields = ['cedula']
    list_filter = ['codigo_masivo']

admin.site.register(ParticipanteDocenteMasivo, ParticipanteDocenteMasivoAdmin)







def construir_proyecto(modeladmin,request,queryset):
    for codigo in queryset:
        participantes = ParticipanteProyectoMasivos.objects.filter(codigo_masivo=codigo)
        redas = Nivel1_Sesion1_REDA.objects.order_by('?')[0:3]

        bibliografia = ""

        for reda in redas:
            bibliografia += reda.portal+"\n"

        x = random.randint(0,2)

        if x == 0:
            previo = Nivel4_Sesion2_1.objects.order_by('?').first().respuesta

        if x == 1:
            previo = Nivel4_Sesion2_2.objects.order_by('?').first().respuesta

        if x == 2:
            previo = Nivel4_Sesion2_2.objects.order_by('?').first().respuesta





        if participantes.count() == 5:

            fields = [  ('Campo 1',participantes[0].nombre.capitalize()),
                        ('Campo de texto 2',participantes[0].celular),
                        ('Campo de texto 3',participantes[0].correo),
                        ('Campo de texto 4',participantes[0].departamento.capitalize()),
                        ('Campo de texto 5',participantes[0].municipio.capitalize()),
                        ('Campo de texto 6',participantes[0].institucion.capitalize()),
                        ('Campo de texto 8',participantes[0].sede.capitalize()),
                        ('Campo de texto 9',participantes[0].dane),
                        ('Campo de texto 10',""),
                        ('Campo de texto 11',""),

                        ('Campo de texto 12',participantes[1].nombre.capitalize()),
                        ('Campo de texto 13',participantes[1].celular),
                        ('Campo de texto 14',participantes[1].correo),
                        ('Campo de texto 15',participantes[1].departamento.capitalize()),
                        ('Campo de texto 16',participantes[1].municipio.capitalize()),
                        ('Campo de texto 17',participantes[1].institucion.capitalize()),
                        ('Campo de texto 18',participantes[1].sede.capitalize()),
                        ('Campo de texto 19',participantes[1].dane),
                        ('Campo de texto 20',""),
                        ('Campo de texto 21',""),

                        ('Campo de texto 22',participantes[2].nombre.capitalize()),
                        ('Campo de texto 23',participantes[2].celular),
                        ('Campo de texto 24',participantes[2].correo),
                        ('Campo de texto 25',participantes[2].departamento.capitalize()),
                        ('Campo de texto 26',participantes[2].municipio.capitalize()),
                        ('Campo de texto 27',participantes[2].institucion.capitalize()),
                        ('Campo de texto 28',participantes[2].sede.capitalize()),
                        ('Campo de texto 29',participantes[2].dane),
                        ('Campo de texto 30',""),
                        ('Campo de texto 31',""),

                        ('Campo de texto 32',participantes[3].nombre.capitalize()),
                        ('Campo de texto 33',participantes[3].celular),
                        ('Campo de texto 34',participantes[3].correo),
                        ('Campo de texto 35',participantes[3].departamento.capitalize()),
                        ('Campo de texto 36',participantes[3].municipio.capitalize()),
                        ('Campo de texto 37',participantes[3].institucion.capitalize()),
                        ('Campo de texto 38',participantes[3].sede.capitalize()),
                        ('Campo de texto 39',participantes[3].dane),
                        ('Campo de texto 40',""),
                        ('Campo de texto 41',""),

                        ('Campo de texto 42',participantes[4].nombre.capitalize()),
                        ('Campo de texto 43',participantes[4].celular),
                        ('Campo de texto 44',participantes[4].correo),
                        ('Campo de texto 45',participantes[4].departamento.capitalize()),
                        ('Campo de texto 46',participantes[4].municipio.capitalize()),
                        ('Campo de texto 47',participantes[4].institucion.capitalize()),
                        ('Campo de texto 48',participantes[4].sede.capitalize()),
                        ('Campo de texto 49',participantes[4].dane),
                        ('Campo de texto 50',""),
                        ('Campo de texto 51',""),

                        ('Campo de texto 52',redas[0].recurso),
                        ('Campo de texto 53',redas[0].portal),
                        ('Campo de texto 54',redas[0].url),

                        ('Campo de texto 55',redas[1].recurso),
                        ('Campo de texto 56',redas[1].portal),
                        ('Campo de texto 57',redas[1].url),

                        ('Campo de texto 153',redas[2].recurso),
                        ('Campo de texto 154',redas[2].portal),
                        ('Campo de texto 155',redas[2].url),

                        ('Campo de texto 84',codigo.nombre_proyecto),
                        ('Campo de texto 91',codigo.definicion_problema),
                        ('Campo de texto 120',codigo.competencia),
                        ('Campo de texto 1012',previo),
                        ('Campo de texto 1038',bibliografia),
                        ('Campo de texto 1048','http://sican.asoandes.org/region/2/cpe/formacion/etic@/docentes/participantes/'),

                        ('Campo de texto 136',Nivel4_Sesion3_2.objects.order_by('?').first().respuesta),
                        ('Campo de texto 137',Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),
                        ('Campo de texto 138',Nivel4_Sesion3_2.objects.order_by('?').first().respuesta),
                        ('Campo de texto 140',Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),
                        ('Campo de texto 139',Nivel4_Sesion3_2.objects.order_by('?').first().respuesta),
                        ('Campo de texto 141',Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),


                        ('Campo de texto 1026',Nivel1_Sesion2_1.objects.order_by('?').first().respuesta+'\n'+Nivel1_Sesion2_2.objects.order_by('?').first().respuesta+'\n'+Nivel1_Sesion2_3.objects.order_by('?').first().respuesta+'\n'+Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),
                        ('Campo de texto 1047','http://sican.asoandes.org/region/2/cpe/formacion/etic@/docentes/participantes/'),
                        ('Campo de texto 1015','http://sican.asoandes.org/region/2/cpe/formacion/etic@/docentes/participantes/'),
                        ]
        if participantes.count() == 4:

            fields = [  ('Campo 1',participantes[0].nombre.capitalize()),
                        ('Campo de texto 2',participantes[0].celular),
                        ('Campo de texto 3',participantes[0].correo),
                        ('Campo de texto 4',participantes[0].departamento.capitalize()),
                        ('Campo de texto 5',participantes[0].municipio.capitalize()),
                        ('Campo de texto 6',participantes[0].institucion.capitalize()),
                        ('Campo de texto 8',participantes[0].sede.capitalize()),
                        ('Campo de texto 9',participantes[0].dane),
                        ('Campo de texto 10',""),
                        ('Campo de texto 11',""),

                        ('Campo de texto 12',participantes[1].nombre.capitalize()),
                        ('Campo de texto 13',participantes[1].celular),
                        ('Campo de texto 14',participantes[1].correo),
                        ('Campo de texto 15',participantes[1].departamento.capitalize()),
                        ('Campo de texto 16',participantes[1].municipio.capitalize()),
                        ('Campo de texto 17',participantes[1].institucion.capitalize()),
                        ('Campo de texto 18',participantes[1].sede.capitalize()),
                        ('Campo de texto 19',participantes[1].dane),
                        ('Campo de texto 20',""),
                        ('Campo de texto 21',""),

                        ('Campo de texto 22',participantes[2].nombre.capitalize()),
                        ('Campo de texto 23',participantes[2].celular),
                        ('Campo de texto 24',participantes[2].correo),
                        ('Campo de texto 25',participantes[2].departamento.capitalize()),
                        ('Campo de texto 26',participantes[2].municipio.capitalize()),
                        ('Campo de texto 27',participantes[2].institucion.capitalize()),
                        ('Campo de texto 28',participantes[2].sede.capitalize()),
                        ('Campo de texto 29',participantes[2].dane),
                        ('Campo de texto 30',""),
                        ('Campo de texto 31',""),

                        ('Campo de texto 32',participantes[3].nombre.capitalize()),
                        ('Campo de texto 33',participantes[3].celular),
                        ('Campo de texto 34',participantes[3].correo),
                        ('Campo de texto 35',participantes[3].departamento.capitalize()),
                        ('Campo de texto 36',participantes[3].municipio.capitalize()),
                        ('Campo de texto 37',participantes[3].institucion.capitalize()),
                        ('Campo de texto 38',participantes[3].sede.capitalize()),
                        ('Campo de texto 39',participantes[3].dane),
                        ('Campo de texto 40',""),
                        ('Campo de texto 41',""),

                        ('Campo de texto 52',redas[0].recurso),
                        ('Campo de texto 53',redas[0].portal),
                        ('Campo de texto 54',redas[0].url),

                        ('Campo de texto 55',redas[1].recurso),
                        ('Campo de texto 56',redas[1].portal),
                        ('Campo de texto 57',redas[1].url),

                        ('Campo de texto 153',redas[2].recurso),
                        ('Campo de texto 154',redas[2].portal),
                        ('Campo de texto 155',redas[2].url),

                        ('Campo de texto 84',codigo.nombre_proyecto),
                        ('Campo de texto 91',codigo.definicion_problema),
                        ('Campo de texto 120',codigo.competencia),
                        ('Campo de texto 1012',previo),
                        ('Campo de texto 1038',bibliografia),

                        ('Campo de texto 136',Nivel4_Sesion3_2.objects.order_by('?').first().respuesta),
                        ('Campo de texto 137',Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),
                        ('Campo de texto 138',Nivel4_Sesion3_2.objects.order_by('?').first().respuesta),
                        ('Campo de texto 140',Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),
                        ('Campo de texto 139',Nivel4_Sesion3_2.objects.order_by('?').first().respuesta),
                        ('Campo de texto 141',Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),

                        ('Campo de texto 1026',Nivel1_Sesion2_1.objects.order_by('?').first().respuesta+'\n'+Nivel1_Sesion2_2.objects.order_by('?').first().respuesta+'\n'+Nivel1_Sesion2_3.objects.order_by('?').first().respuesta+'\n'+Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),
                        ('Campo de texto 1047','http://sican.asoandes.org/region/2/cpe/formacion/etic@/docentes/participantes/'),
                        ('Campo de texto 1048','http://sican.asoandes.org/region/2/cpe/formacion/etic@/docentes/participantes/'),
                        ('Campo de texto 1015','http://sican.asoandes.org/region/2/cpe/formacion/etic@/docentes/participantes/'),
                        ]

        if participantes.count() == 3:

            fields = [  ('Campo 1',participantes[0].nombre.capitalize()),
                        ('Campo de texto 2',participantes[0].celular),
                        ('Campo de texto 3',participantes[0].correo),
                        ('Campo de texto 4',participantes[0].departamento.capitalize()),
                        ('Campo de texto 5',participantes[0].municipio.capitalize()),
                        ('Campo de texto 6',participantes[0].institucion.capitalize()),
                        ('Campo de texto 8',participantes[0].sede.capitalize()),
                        ('Campo de texto 9',participantes[0].dane),
                        ('Campo de texto 10',""),
                        ('Campo de texto 11',""),

                        ('Campo de texto 12',participantes[1].nombre.capitalize()),
                        ('Campo de texto 13',participantes[1].celular),
                        ('Campo de texto 14',participantes[1].correo),
                        ('Campo de texto 15',participantes[1].departamento.capitalize()),
                        ('Campo de texto 16',participantes[1].municipio.capitalize()),
                        ('Campo de texto 17',participantes[1].institucion.capitalize()),
                        ('Campo de texto 18',participantes[1].sede.capitalize()),
                        ('Campo de texto 19',participantes[1].dane),
                        ('Campo de texto 20',""),
                        ('Campo de texto 21',""),

                        ('Campo de texto 22',participantes[2].nombre.capitalize()),
                        ('Campo de texto 23',participantes[2].celular),
                        ('Campo de texto 24',participantes[2].correo),
                        ('Campo de texto 25',participantes[2].departamento.capitalize()),
                        ('Campo de texto 26',participantes[2].municipio.capitalize()),
                        ('Campo de texto 27',participantes[2].institucion.capitalize()),
                        ('Campo de texto 28',participantes[2].sede.capitalize()),
                        ('Campo de texto 29',participantes[2].dane),
                        ('Campo de texto 30',""),
                        ('Campo de texto 31',""),

                        ('Campo de texto 52',redas[0].recurso),
                        ('Campo de texto 53',redas[0].portal),
                        ('Campo de texto 54',redas[0].url),

                        ('Campo de texto 55',redas[1].recurso),
                        ('Campo de texto 56',redas[1].portal),
                        ('Campo de texto 57',redas[1].url),

                        ('Campo de texto 153',redas[2].recurso),
                        ('Campo de texto 154',redas[2].portal),
                        ('Campo de texto 155',redas[2].url),

                        ('Campo de texto 84',codigo.nombre_proyecto),
                        ('Campo de texto 91',codigo.definicion_problema),
                        ('Campo de texto 120',codigo.competencia),
                        ('Campo de texto 1012',previo),
                        ('Campo de texto 1038',bibliografia),

                        ('Campo de texto 136',Nivel4_Sesion3_2.objects.order_by('?').first().respuesta),
                        ('Campo de texto 137',Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),
                        ('Campo de texto 138',Nivel4_Sesion3_2.objects.order_by('?').first().respuesta),
                        ('Campo de texto 140',Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),
                        ('Campo de texto 139',Nivel4_Sesion3_2.objects.order_by('?').first().respuesta),
                        ('Campo de texto 141',Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),

                        ('Campo de texto 1026',Nivel1_Sesion2_1.objects.order_by('?').first().respuesta+'\n'+Nivel1_Sesion2_2.objects.order_by('?').first().respuesta+'\n'+Nivel1_Sesion2_3.objects.order_by('?').first().respuesta+'\n'+Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),
                        ('Campo de texto 1047','http://sican.asoandes.org/region/2/cpe/formacion/etic@/docentes/participantes/'),
                        ('Campo de texto 1048','http://sican.asoandes.org/region/2/cpe/formacion/etic@/docentes/participantes/'),
                        ('Campo de texto 1015','http://sican.asoandes.org/region/2/cpe/formacion/etic@/docentes/participantes/'),
                        ]


        if participantes.count() == 2:

            fields = [  ('Campo 1',participantes[0].nombre.capitalize()),
                        ('Campo de texto 2',participantes[0].celular),
                        ('Campo de texto 3',participantes[0].correo),
                        ('Campo de texto 4',participantes[0].departamento.capitalize()),
                        ('Campo de texto 5',participantes[0].municipio.capitalize()),
                        ('Campo de texto 6',participantes[0].institucion.capitalize()),
                        ('Campo de texto 8',participantes[0].sede.capitalize()),
                        ('Campo de texto 9',participantes[0].dane),
                        ('Campo de texto 10',""),
                        ('Campo de texto 11',""),

                        ('Campo de texto 12',participantes[1].nombre.capitalize()),
                        ('Campo de texto 13',participantes[1].celular),
                        ('Campo de texto 14',participantes[1].correo),
                        ('Campo de texto 15',participantes[1].departamento.capitalize()),
                        ('Campo de texto 16',participantes[1].municipio.capitalize()),
                        ('Campo de texto 17',participantes[1].institucion.capitalize()),
                        ('Campo de texto 18',participantes[1].sede.capitalize()),
                        ('Campo de texto 19',participantes[1].dane),
                        ('Campo de texto 20',""),
                        ('Campo de texto 21',""),

                        ('Campo de texto 52',redas[0].recurso),
                        ('Campo de texto 53',redas[0].portal),
                        ('Campo de texto 54',redas[0].url),

                        ('Campo de texto 55',redas[1].recurso),
                        ('Campo de texto 56',redas[1].portal),
                        ('Campo de texto 57',redas[1].url),

                        ('Campo de texto 153',redas[2].recurso),
                        ('Campo de texto 154',redas[2].portal),
                        ('Campo de texto 155',redas[2].url),

                        ('Campo de texto 84',codigo.nombre_proyecto),
                        ('Campo de texto 91',codigo.definicion_problema),
                        ('Campo de texto 120',codigo.competencia),
                        ('Campo de texto 1012',previo),
                        ('Campo de texto 1038',bibliografia),

                        ('Campo de texto 136',Nivel4_Sesion3_2.objects.order_by('?').first().respuesta),
                        ('Campo de texto 137',Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),
                        ('Campo de texto 138',Nivel4_Sesion3_2.objects.order_by('?').first().respuesta),
                        ('Campo de texto 140',Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),
                        ('Campo de texto 139',Nivel4_Sesion3_2.objects.order_by('?').first().respuesta),
                        ('Campo de texto 141',Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),

                        ('Campo de texto 1026',Nivel1_Sesion2_1.objects.order_by('?').first().respuesta+'\n'+Nivel1_Sesion2_2.objects.order_by('?').first().respuesta+'\n'+Nivel1_Sesion2_3.objects.order_by('?').first().respuesta+'\n'+Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),
                        ('Campo de texto 1047','http://sican.asoandes.org/region/2/cpe/formacion/etic@/docentes/participantes/'),
                        ('Campo de texto 1048','http://sican.asoandes.org/region/2/cpe/formacion/etic@/docentes/participantes/'),
                        ('Campo de texto 1015','http://sican.asoandes.org/region/2/cpe/formacion/etic@/docentes/participantes/'),
                        ]

        if participantes.count() == 1:

            fields = [  ('Campo 1',participantes[0].nombre.capitalize()),
                        ('Campo de texto 2',participantes[0].celular),
                        ('Campo de texto 3',participantes[0].correo),
                        ('Campo de texto 4',participantes[0].departamento.capitalize()),
                        ('Campo de texto 5',participantes[0].municipio.capitalize()),
                        ('Campo de texto 6',participantes[0].institucion.capitalize()),
                        ('Campo de texto 8',participantes[0].sede.capitalize()),
                        ('Campo de texto 9',participantes[0].dane),
                        ('Campo de texto 10',""),
                        ('Campo de texto 11',""),

                        ('Campo de texto 52',redas[0].recurso),
                        ('Campo de texto 53',redas[0].portal),
                        ('Campo de texto 54',redas[0].url),

                        ('Campo de texto 55',redas[1].recurso),
                        ('Campo de texto 56',redas[1].portal),
                        ('Campo de texto 57',redas[1].url),

                        ('Campo de texto 153',redas[2].recurso),
                        ('Campo de texto 154',redas[2].portal),
                        ('Campo de texto 155',redas[2].url),

                        ('Campo de texto 84',codigo.nombre_proyecto),
                        ('Campo de texto 91',codigo.definicion_problema),
                        ('Campo de texto 120',codigo.competencia),
                        ('Campo de texto 1012',previo),
                        ('Campo de texto 1038',bibliografia),

                        ('Campo de texto 136',Nivel4_Sesion3_2.objects.order_by('?').first().respuesta),
                        ('Campo de texto 137',Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),
                        ('Campo de texto 138',Nivel4_Sesion3_2.objects.order_by('?').first().respuesta),
                        ('Campo de texto 140',Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),
                        ('Campo de texto 139',Nivel4_Sesion3_2.objects.order_by('?').first().respuesta),
                        ('Campo de texto 141',Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),

                        ('Campo de texto 1026',Nivel1_Sesion2_1.objects.order_by('?').first().respuesta+'\n'+Nivel1_Sesion2_2.objects.order_by('?').first().respuesta+'\n'+Nivel1_Sesion2_3.objects.order_by('?').first().respuesta+'\n'+Nivel1_Sesion2_4.objects.order_by('?').first().respuesta),
                        ('Campo de texto 1047','http://sican.asoandes.org/region/2/cpe/formacion/etic@/docentes/participantes/'),
                        ('Campo de texto 1048','http://sican.asoandes.org/region/2/cpe/formacion/etic@/docentes/participantes/'),
                        ('Campo de texto 1015','http://sican.asoandes.org/region/2/cpe/formacion/etic@/docentes/participantes/'),
                        ]

        fdf = forge_fdf("",fields,[],[],[])
        fdf_file = open("C:\\Temp\\Proyectos\\"+str(codigo.codigo)+".fdf","wb")
        fdf_file.write(fdf)
        fdf_file.close()
        os.system('pdftk C:\\Temp\\Proyectos\\'+str(codigo.codigo)+'.pdf fill_form C:\\Temp\\Proyectos\\'+str(codigo.codigo)+'.fdf output C:\\Temp\\ProyectosListos\\'+str(codigo.codigo)+'.pdf')

construir_proyecto.short_description = "Construir Proyecto"







class CodigoMasivo_ProyectossAdmin(admin.ModelAdmin):
    list_display = ['codigo']
    ordering = ['codigo']
    list_filter = ['region','generado']
    actions = [construir_proyecto]
admin.site.register(CodigoMasivo_Proyectoss,CodigoMasivo_ProyectossAdmin)
admin.site.register(ParticipanteProyectoMasivos)














def generar_listas(modeladmin,request,queryset):
    pythoncom.CoInitializeEx(pythoncom.COINIT_MULTITHREADED)
    xl = win32com.client.dynamic.Dispatch('Excel.Application')
    xl.DisplayAlerts = True
    xl.Visible = 0

    lista = xl.Workbooks.Open(settings.STATICFILES_DIRS[0]+'/formatos/Lista Padres.xlsx')

    i = 2

    for codigo_masivo in queryset:
        participantes = ParticipanteEscuelaTicTrucho.objects.filter(codigo_masivo__id=codigo_masivo.id)
        if len(participantes) > 0:
            for sesion_numero in range(1,3):
                fila = 0
                if sesion_numero == 1:
                    lista.Worksheets("Sesion").Copy(lista.Worksheets(lista.Worksheets.Count))
                    sesion = lista.Worksheets('Sesion ('+str(i)+')')
                    i += 1
                    sesion.Cells(5,3).Value = "Sesión 1:______X_______".encode("latin1")
                    sesion.Cells(5,5).Value = "Sesión 2:______________".encode("latin1")
                    departamento = participantes[0].departamento
                    municipio = participantes[0].municipio
                    formador = participantes[0].formador.nombre
                    sesion.Cells(6,1).Value = "Departamento: "+departamento.encode("latin1")+"                         "+"Municipio: "+municipio.encode("latin1")
                    sesion.Cells(5,7).Value = "Formador: "+formador.encode("latin1")
                    for participante in participantes:
                        sesion.Cells(10+fila,2).Value = participante.nombres.encode("latin1")
                        sesion.Cells(10+fila,3).Value = participante.apellidos.encode("latin1")
                        sesion.Cells(10+fila,4).Value = participante.cedula
                        #sesion.Cells(10+fila,5).Value = participante.correo
                        sesion.Cells(10+fila,6).Value = participante.telefono
                        sesion.Cells(10+fila,9).Value = "X"
                        sesion.Cells(10+fila,10).Value = "X"
                        sesion.Cells(10+fila,11).Value = "X"
                        fila += 1
                if sesion_numero == 2:
                    lista.Worksheets("Sesion").Copy(lista.Worksheets(lista.Worksheets.Count))
                    sesion = lista.Worksheets('Sesion ('+str(i)+')')
                    i += 1
                    sesion.Cells(5,3).Value = "Sesión 1:______________".encode("latin1")
                    sesion.Cells(5,5).Value = "Sesión 2:______X_______".encode("latin1")
                    departamento = participantes[0].departamento
                    municipio = participantes[0].municipio
                    formador = participantes[0].formador.nombre
                    sesion.Cells(6,1).Value = "Departamento: "+departamento.encode("latin1")+"                         "+"Municipio: "+municipio.encode("latin1")
                    sesion.Cells(5,7).Value = "Formador: "+formador.encode("latin1")
                    for participante in participantes:
                        sesion.Cells(10+fila,2).Value = participante.nombres.encode("latin1")
                        sesion.Cells(10+fila,3).Value = participante.apellidos.encode("latin1")
                        sesion.Cells(10+fila,4).Value = participante.cedula
                        #sesion.Cells(10+fila,5).Value = participante.correo
                        sesion.Cells(10+fila,6).Value = participante.telefono
                        sesion.Cells(10+fila,9).Value = "X"
                        sesion.Cells(10+fila,10).Value = "X"
                        sesion.Cells(10+fila,11).Value = "X"
                        sesion.Cells(10+fila,12).Value = "X"
                        fila += 1

    lista.Worksheets('Sesion').Delete()
    lista.ExportAsFixedFormat(0,settings.MEDIA_ROOT+'/Listados Escuela Tic/Padres.pdf')
    lista.Close(SaveChanges=False)

    xl.Quit()
    pythoncom.CoUninitialize()

    for codigo_masivo in queryset:
        codigo_masivo.generado = True
        codigo_masivo.save()

    return HttpResponseRedirect('/media/Listados Escuela Tic/Padres.pdf')
generar_listas.short_description = "Generar Listas"

def copiar_participantes(modeladmin,request,queryset):
    for participante_trucho in ParticipanteEscuelaTicTrucho.objects.all():
        if ParticipanteEscuelaTic.objects.filter(cedula=participante_trucho.cedula).count() == 0:
            nuevo = ParticipanteEscuelaTic()
            nuevo.formador = participante_trucho.formador
            nuevo.grupo = participante_trucho.grupo
            nuevo.institucion = participante_trucho.institucion
            nuevo.nombres = participante_trucho.nombres
            nuevo.apellidos = participante_trucho.apellidos
            nuevo.cedula = participante_trucho.cedula
            nuevo.genero = participante_trucho.genero
            nuevo.nivel_educativo = participante_trucho.nivel_educativo
            nuevo.telefono = participante_trucho.telefono
            nuevo.correo = participante_trucho.correo
            nuevo.poblacion = participante_trucho.poblacion
            nuevo.codigo_anspe = participante_trucho.codigo_anspe
            nuevo.tipo_proyecto = participante_trucho.tipo_proyecto
            nuevo.grupo_conformacion = participante_trucho.grupo_conformacion
            nuevo.save()
copiar_participantes.short_description = "Copiar Participantes"

class CodigoMasivoAdmin(admin.ModelAdmin):
    list_display = ['id','generado']
    list_filter = ['generado']
    ordering = ['id']
    actions = [generar_listas,copiar_participantes]
admin.site.register(CodigoMasivo,CodigoMasivoAdmin)

def generar_virtual_1(modeladmin,request,queryset):
    evidencia_docentes = EvidenciaDocentes.objects.filter(entregable__id=11,soporte=None)
    for evidencia_docente in evidencia_docentes:
        nuevo = SoporteEntregableDocente(grupo=evidencia_docente.participante.grupo,entregable=EntregableDocentes.objects.get(id=11))
        redas = Nivel1_Sesion1_REDA.objects.order_by('?')[:5].values_list('id',flat=True)
        redas = list(redas)
        fields = [('Campo de texto 202',Nivel1_Sesion1_1.objects.order_by('?').first()),
                    ('Campo de texto 203',Nivel1_Sesion1_2.objects.order_by('?').first()),
                    ('Campo de texto 204',Nivel1_Sesion1_3.objects.order_by('?').first()),
                    ('Campo de texto 205',Nivel1_Sesion1_4.objects.order_by('?').first()),
                    ('Campo de texto 206',Nivel1_Sesion1_5.objects.order_by('?').first()),
                    ('Campo de texto 207',Nivel1_Sesion1_6.objects.order_by('?').first()),
                    ('Campo de texto 208',Nivel1_Sesion1_7.objects.order_by('?').first()),
                    ('Campo de texto 209',Nivel1_Sesion1_8.objects.order_by('?').first()),
                    ('Campo de texto 2010',Nivel1_Sesion1_9.objects.order_by('?').first()),
                    ('Campo de texto 2011',Nivel1_Sesion1_10.objects.order_by('?').first()),
                    ('Campo de texto 2012',Nivel1_Sesion1_11.objects.order_by('?').first()),
                    ('Campo de texto 2013',Nivel1_Sesion1_12.objects.order_by('?').first()),

                  ('a1','Elección'+str(random.randrange(1,5))),
                  ('a2','Elección'+str(random.randrange(1,5))),
                  ('a3','Elección'+str(random.randrange(1,5))),
                  ('a4','Elección'+str(random.randrange(1,5))),
                  ('a5','Elección'+str(random.randrange(1,5))),
                  ('a6','Elección'+str(random.randrange(1,5))),
                  ('a7','Elección'+str(random.randrange(1,5))),
                  ('a8','Elección'+str(random.randrange(1,5))),
                  ('a9','Elección'+str(random.randrange(1,5))),
                  ('a10','Elección'+str(random.randrange(1,5))),
                  ('a11','Elección'+str(random.randrange(1,5))),
                  ('a12','Elección'+str(random.randrange(1,5))),
                  ('a13','Elección'+str(random.randrange(1,5))),
                  ('a14','Elección'+str(random.randrange(1,5))),
                  ('a15','Elección'+str(random.randrange(1,5))),
                  ('a16','Elección'+str(random.randrange(1,5))),
                  ('a17','Elección'+str(random.randrange(1,5))),
                  ('a18','Elección'+str(random.randrange(1,5))),
                  ('a19','Elección'+str(random.randrange(1,5))),
                  ('a20','Elección'+str(random.randrange(1,5))),
                  ('Campo de texto 2031','1'),
                  ('Campo de texto 2026',Nivel1_Sesion1_REDA.objects.get(id=redas[0]).recurso),
                  ('Campo de texto 2016',Nivel1_Sesion1_REDA.objects.get(id=redas[0]).portal),
                  ('Campo de texto 2021',Nivel1_Sesion1_REDA.objects.get(id=redas[0]).url),

                  ('Campo de texto 2032','2'),
                  ('Campo de texto 2027',Nivel1_Sesion1_REDA.objects.get(id=redas[1]).recurso),
                  ('Campo de texto 2017',Nivel1_Sesion1_REDA.objects.get(id=redas[1]).portal),
                  ('Campo de texto 2022',Nivel1_Sesion1_REDA.objects.get(id=redas[1]).url),

                  ('Campo de texto 2034','3'),
                  ('Campo de texto 2029',Nivel1_Sesion1_REDA.objects.get(id=redas[2]).recurso),
                  ('Campo de texto 2018',Nivel1_Sesion1_REDA.objects.get(id=redas[2]).portal),
                  ('Campo de texto 2024',Nivel1_Sesion1_REDA.objects.get(id=redas[2]).url),

                  ('Campo de texto 2033','4'),
                  ('Campo de texto 2028',Nivel1_Sesion1_REDA.objects.get(id=redas[3]).recurso),
                  ('Campo de texto 2019',Nivel1_Sesion1_REDA.objects.get(id=redas[3]).portal),
                  ('Campo de texto 2023',Nivel1_Sesion1_REDA.objects.get(id=redas[3]).url),

                  ('Campo de texto 2035','5'),
                  ('Campo de texto 2030',Nivel1_Sesion1_REDA.objects.get(id=redas[4]).recurso),
                  ('Campo de texto 2020',Nivel1_Sesion1_REDA.objects.get(id=redas[4]).portal),
                  ('Campo de texto 2025',Nivel1_Sesion1_REDA.objects.get(id=redas[4]).url),
                    ]
        fdf = forge_fdf("",fields,[],[],[])
        fdf_file = open("C:\\Temp\\data.fdf","wb")
        fdf_file.write(fdf)
        fdf_file.close()
        os.system('pdftk C:\\Temp\\sesion1_plantilla.pdf fill_form C:\\Temp\\data.fdf output C:\\Temp\\sesion1.pdf flatten')
        nuevo.soporte = File(open("C://Temp//sesion1.pdf", 'rb'))
        nuevo.save()
        evidencia_docente.soporte = nuevo
        evidencia_docente.save()
generar_virtual_1.short_description = "Generar Actividad Virtual 1"

def generar_sesion5_nive1(modeladmin,request,queryset):
    evidencia_docentes = EvidenciaDocentes.objects.filter(entregable__id=9,soporte=None)
    for evidencia_docente in evidencia_docentes:
        nuevo = SoporteEntregableDocente(grupo=evidencia_docente.participante.grupo,entregable=EntregableDocentes.objects.get(id=9))
        fields = [('N51','Elección'+str(random.randrange(1,3))),
                  ('N52','Elección'+str(random.randrange(1,3))),
                  ('N53','Elección'+str(random.randrange(1,3))),
                  ('N54','Elección'+str(random.randrange(1,6))),
                  ('N55','Elección'+str(random.randrange(1,3))),
                  ('N56','Elección'+str(random.randrange(1,3))),
                  ('N58','Elección'+str(random.randrange(1,3))),
                 ]
        fdf = forge_fdf("",fields,[],[],[])
        fdf_file = open("C:\\Temp\\datas5n1.fdf","wb")
        fdf_file.write(fdf)
        fdf_file.close()
        os.system('pdftk C:\\Temp\\sesion5_n1_plantilla.pdf fill_form C:\\Temp\\datas5n1.fdf output C:\\Temp\\sesion5.pdf flatten')
        nuevo.soporte = File(open("C://Temp//sesion5.pdf", 'rb'))
        nuevo.save()
        evidencia_docente.soporte = nuevo
        evidencia_docente.save()
generar_sesion5_nive1.short_description = "Generar Sesion 5 - Nivel 1"



def generar_sesion5_nive1(modeladmin,request,queryset):
    evidencia_docentes = EvidenciaDocentes.objects.filter(entregable__id=9,soporte=None)
    for evidencia_docente in evidencia_docentes:
        nuevo = SoporteEntregableDocente(grupo=evidencia_docente.participante.grupo,entregable=EntregableDocentes.objects.get(id=9))
        fields = [('N51','Elección'+str(random.randrange(1,3))),
                  ('N52','Elección'+str(random.randrange(1,3))),
                  ('N53','Elección'+str(random.randrange(1,3))),
                  ('N54','Elección'+str(random.randrange(1,6))),
                  ('N55','Elección'+str(random.randrange(1,3))),
                  ('N56','Elección'+str(random.randrange(1,3))),
                  ('N58','Elección'+str(random.randrange(1,3))),
                 ]
        fdf = forge_fdf("",fields,[],[],[])
        fdf_file = open("C:\\Temp\\datas5n1.fdf","wb")
        fdf_file.write(fdf)
        fdf_file.close()
        os.system('pdftk C:\\Temp\\sesion5_n1_plantilla.pdf fill_form C:\\Temp\\datas5n1.fdf output C:\\Temp\\sesion5.pdf flatten')
        nuevo.soporte = File(open("C://Temp//sesion5.pdf", 'rb'))
        nuevo.save()
        evidencia_docente.soporte = nuevo
        evidencia_docente.save()
generar_sesion5_nive1.short_description = "Generar Sesion 5 - Nivel 1"



class Nivel1_Sesion1_1Admin(admin.ModelAdmin):
    list_display = ['respuesta']
    actions = [generar_virtual_1,generar_sesion5_nive1,reporte_formadores,reporte_formadores_tipo2,reporte_docentes]



def generar_nivel1_sesion2(modeladmin,request,queryset):
    evidencia_docentes = EvidenciaDocentes.objects.filter(entregable__id=13,soporte=None)
    for evidencia_docente in evidencia_docentes:
        nuevo = SoporteEntregableDocente(grupo=evidencia_docente.participante.grupo,entregable=EntregableDocentes.objects.get(id=13))
        fields = [('Campo de texto 254','http://sican.asoandes.org:8000/index.php/apps/files/?dir=%2FNivel%201%2FSesion%202'),
                  ('Campo de texto 255',Nivel1_Sesion2.objects.order_by('?').first()),

                 ]
        fdf = forge_fdf("",fields,[],[],[])
        fdf_file = open("C:\\Temp\\datas2n1.fdf","wb")
        fdf_file.write(fdf)
        fdf_file.close()
        os.system('pdftk C:\\Temp\\sesion2_n1_plantilla.pdf fill_form C:\\Temp\\datas2n1.fdf output C:\\Temp\\sesion2.pdf flatten')
        nuevo.soporte = File(open("C://Temp//sesion2.pdf", 'rb'))
        nuevo.save()
        evidencia_docente.soporte = nuevo
        evidencia_docente.save()
generar_nivel1_sesion2.short_description = "Generar Sesion 2 - Nivel 1"

class Nivel1_Sesion2Admin(admin.ModelAdmin):
    list_display = ['respuesta']
    actions = [generar_nivel1_sesion2]
admin.site.register(Nivel1_Sesion2,Nivel1_Sesion2Admin)


def generar_nivel1_sesion3(modeladmin,request,queryset):
    evidencia_docentes = EvidenciaDocentes.objects.filter(entregable__id=15,soporte=None)
    for evidencia_docente in evidencia_docentes:
        nuevo = SoporteEntregableDocente(grupo=evidencia_docente.participante.grupo,entregable=EntregableDocentes.objects.get(id=15))
        fields = [('Campo de texto 256',Nivel1_Sesion3.objects.order_by('?').first()),
                 ]
        fdf = forge_fdf("",fields,[],[],[])
        fdf_file = open("C:\\Temp\\datas3n1.fdf","wb")
        fdf_file.write(fdf)
        fdf_file.close()
        os.system('pdftk C:\\Temp\\sesion3_n1_plantilla.pdf fill_form C:\\Temp\\datas3n1.fdf output C:\\Temp\\sesion3.pdf flatten')
        nuevo.soporte = File(open("C://Temp//sesion3.pdf", 'rb'))
        nuevo.save()
        evidencia_docente.soporte = nuevo
        evidencia_docente.save()
generar_nivel1_sesion3.short_description = "Generar Sesion 3 - Nivel 1"

def generar_nivel2_sesion2(modeladmin,request,queryset):
    evidencia_docentes = EvidenciaDocentes.objects.filter(entregable__id=27,soporte=None)
    for evidencia_docente in evidencia_docentes:
        nuevo = SoporteEntregableDocente(grupo=evidencia_docente.participante.grupo,entregable=EntregableDocentes.objects.get(id=27))
        fields = [('R1-1','Elección'+str(random.randrange(1,4))),
                  ('R1-2','Elección'+str(random.randrange(1,4))),
                  ('R1-3','Elección'+str(random.randrange(1,4))),
                  ('R1-4','Elección'+str(random.randrange(1,4))),
                  ('R1-5','Elección'+str(random.randrange(1,4))),
                  ('R1-6','Elección'+str(random.randrange(1,4))),
                  ('R1-7','Elección'+str(random.randrange(1,3))),
                  ('R1-8','Elección'+str(random.randrange(1,3))),
                  ('R1-9','Elección'+str(random.randrange(1,3))),
                  ('R1-10','Elección'+str(random.randrange(1,3))),
        ]
        fdf = forge_fdf("",fields,[],[],[])
        fdf_file = open("C:\\Temp\\datas2n2.fdf","wb")
        fdf_file.write(fdf)
        fdf_file.close()
        os.system('pdftk C:\\Temp\\sesion2_n2_plantilla.pdf fill_form C:\\Temp\\datas2n2.fdf output C:\\Temp\\sesion2_2.pdf flatten')
        nuevo.soporte = File(open("C://Temp//sesion2_2.pdf", 'rb'))
        nuevo.save()
        evidencia_docente.soporte = nuevo
        evidencia_docente.save()
generar_nivel2_sesion2.short_description = "Generar Sesion 2 - Nivel 2"

def generar_nivel3_1_sesion3(modeladmin,request,queryset):
    evidencia_docentes = EvidenciaDocentes.objects.filter(entregable__id=39,soporte=None)
    for evidencia_docente in evidencia_docentes:
        nuevo = SoporteEntregableDocente(grupo=evidencia_docente.participante.grupo,entregable=EntregableDocentes.objects.get(id=39))
        fields = [('Campo de texto 3024',evidencia_docente.participante.nombres+ " "+evidencia_docente.participante.apellidos),
                  ('N3s3-1','Elección'+str(random.randrange(1,4))),
                  ('N3s3-2','Elección'+str(random.randrange(1,4))),
                  ('N3s3-3','Elección'+str(random.randrange(1,4))),
                  ('N3s3-4','Elección'+str(random.randrange(1,4))),
                  ('N3s3-5','Elección'+str(random.randrange(1,4))),
                  ('N3s3-6','Elección'+str(random.randrange(1,4))),
                  ('N3s3-7','Elección'+str(random.randrange(1,4))),
                  ('N3s3-8','Elección'+str(random.randrange(1,4))),
        ]
        fdf = forge_fdf("",fields,[],[],[])
        fdf_file = open("C:\\Temp\\datas3_1n3.fdf","wb")
        fdf_file.write(fdf)
        fdf_file.close()
        os.system('pdftk C:\\Temp\\sesion3_1_n3_plantilla.pdf fill_form C:\\Temp\\datas3_1n3.fdf output C:\\Temp\\sesion3_1_3.pdf flatten')
        nuevo.soporte = File(open("C://Temp//sesion3_1_3.pdf", 'rb'))
        nuevo.save()
        evidencia_docente.soporte = nuevo
        evidencia_docente.save()
generar_nivel3_1_sesion3.short_description = "Generar Sesion 3.1 - Nivel 3"



def matriz_escuela_tic(modeladmin,request,queryset):

    data = ParticipanteEscuelaTicTrucho.objects.all()
    chunks=[data[x:x+35000] for x in xrange(0, data.count(), 35000)]
    i = 0
    for chunk in chunks:
        i += 1
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Matriz Escuela Tic '+unicode(i)+'.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Matriz Escuela Tic"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Matriz Escuela Tic'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['Region',30]),
                   tuple(['Departamento',30]),
                   tuple(['Municipio',30]),
                   tuple(['Institucion',30]),
                   tuple(['Codigo del Grupo',30]),
                   tuple(['Nombre del Formador',30]),
                   tuple(['Cedula del Formador',30]),
                   tuple(['Nombres',30]),
                   tuple(['Apellidos',30]),
                   tuple(['Cedula',30]),
                   tuple(['Genero',30]),
                   tuple(['Nivel Educativo',30]),
                   tuple(['Telefono',30]),
                   tuple(['Correo',30]),
                   tuple(['Tipo Poblacion',30]),
                   tuple(['Codigo Proyecto',30]),
                   tuple(['Tipo Proyecto',30]),
                   tuple(['Grupo Conformacion',30]),
                   tuple(['Sesion 1',30]),
                   tuple(['Sesion 2',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        for participante in chunk:
                row_num += 1
                row = [
                    participante.formador.region.nombre,
                    participante.grupo.municipio.departamento.nombre,
                    participante.grupo.municipio.nombre,
                    participante.institucion,
                    participante.grupo.nombre,
                    participante.formador.nombre,
                    participante.formador.cedula,
                    participante.nombres,
                    participante.apellidos,
                    participante.cedula,
                    participante.genero,
                    participante.nivel_educativo,
                    participante.telefono,
                    participante.correo,
                    participante.poblacion,
                    participante.codigo_anspe,
                    participante.tipo_proyecto,
                    participante.grupo_conformacion,
                    unicode(EvidenciaEscuelaTic.objects.filter(participante__id=participante.id).get(entregable__id=5).soporte),
                    unicode(EvidenciaEscuelaTic.objects.filter(participante__id=participante.id).get(entregable__id=9).soporte)
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(settings.MEDIA_ROOT+'/Matriz Padres/Matriz'+unicode(i)+'.xlsx')
    return HttpResponseRedirect('/media/Matriz Padres/Matriz.xlsx')
matriz_escuela_tic.short_description = "Matriz Escuela Tic"

class Nivel1_Sesion3Admin(admin.ModelAdmin):
    list_display = ['respuesta']
    actions = [generar_nivel1_sesion3,generar_nivel2_sesion2,generar_nivel3_1_sesion3,matriz_escuela_tic]
admin.site.register(Nivel1_Sesion3,Nivel1_Sesion3Admin)


def generar_nivel4_1(modeladmin,request,queryset):
    evidencia_docentes = EvidenciaDocentes.objects.filter(entregable__id=51,soporte=None)
    for evidencia_docente in evidencia_docentes:
        nuevo = SoporteEntregableDocente(grupo=evidencia_docente.participante.grupo,entregable=EntregableDocentes.objects.get(id=51))
        fields = [('Campo de texto 3037',Nivel4_Sesion1_1.objects.all().order_by('?').first()),
                  ('Campo de texto 3038',Nivel4_Sesion1_2.objects.all().order_by('?').first())
        ]
        fdf = forge_fdf("",fields,[],[],[])
        fdf_file = open("C:\\Temp\\datas1_n4.fdf","wb")
        fdf_file.write(fdf)
        fdf_file.close()
        os.system('pdftk C:\\Temp\\sesion1_n4_plantilla.pdf fill_form C:\\Temp\\datas1_n4.fdf output C:\\Temp\\sesion1_4.pdf flatten')
        nuevo.soporte = File(open("C://Temp//sesion1_4.pdf", 'rb'))
        nuevo.save()
        evidencia_docente.soporte = nuevo
        evidencia_docente.save()
generar_nivel4_1.short_description = "Generar Sesion 1 - Nivel 4"

class Nivel4_Sesion1_1Admin(admin.ModelAdmin):
    list_display = ['respuesta']
    actions = [generar_nivel4_1]
admin.site.register(Nivel4_Sesion1_1,Nivel4_Sesion1_1Admin)

class Nivel4_Sesion1_2Admin(admin.ModelAdmin):
    list_display = ['respuesta']
    actions = []
admin.site.register(Nivel4_Sesion1_2,Nivel4_Sesion1_2Admin)





def generar_nivel4_2(modeladmin,request,queryset):
    evidencia_docentes = EvidenciaDocentes.objects.filter(entregable__id=53,soporte=None)
    for evidencia_docente in evidencia_docentes:
        nuevo = SoporteEntregableDocente(grupo=evidencia_docente.participante.grupo,entregable=EntregableDocentes.objects.get(id=53))
        fields = [('Campo de texto 3039',Nivel4_Sesion2_1.objects.all().order_by('?').first()),
                  ('Campo de texto 3040',Nivel4_Sesion2_2.objects.all().order_by('?').first()),
                  ('Campo de texto 3041',Nivel4_Sesion2_3.objects.all().order_by('?').first())
        ]
        fdf = forge_fdf("",fields,[],[],[])
        fdf_file = open("C:\\Temp\\datas2_n4.fdf","wb")
        fdf_file.write(fdf)
        fdf_file.close()
        os.system('pdftk C:\\Temp\\sesion2_n4_plantilla.pdf fill_form C:\\Temp\\datas2_n4.fdf output C:\\Temp\\sesion2_4.pdf flatten')
        nuevo.soporte = File(open("C://Temp//sesion2_4.pdf", 'rb'))
        nuevo.save()
        evidencia_docente.soporte = nuevo
        evidencia_docente.save()
generar_nivel4_2.short_description = "Generar Sesion 2 - Nivel 4"

class Nivel4_Sesion2_1Admin(admin.ModelAdmin):
    list_display = ['respuesta']
    actions = [generar_nivel4_2]
admin.site.register(Nivel4_Sesion2_1,Nivel4_Sesion2_1Admin)

class Nivel4_Sesion2_2Admin(admin.ModelAdmin):
    list_display = ['respuesta']
    actions = []
admin.site.register(Nivel4_Sesion2_2,Nivel4_Sesion2_2Admin)

class Nivel4_Sesion2_3Admin(admin.ModelAdmin):
    list_display = ['respuesta']
    actions = []
admin.site.register(Nivel4_Sesion2_3,Nivel4_Sesion2_3Admin)




def generar_nivel4_3(modeladmin,request,queryset):
    evidencia_docentes = EvidenciaDocentes.objects.filter(entregable__id=55,soporte=None)
    for evidencia_docente in evidencia_docentes:
        nuevo = SoporteEntregableDocente(grupo=evidencia_docente.participante.grupo,entregable=EntregableDocentes.objects.get(id=55))
        fields = [('Campo de texto 3042',Nivel4_Sesion3_1.objects.all().order_by('?').first()),
                  ('Campo de texto 3043',Nivel4_Sesion3_2.objects.all().order_by('?').first()),
                  ('Campo de texto 3044',Nivel4_Sesion3_3.objects.all().order_by('?').first()),
                  ('Campo de texto 3045',Nivel4_Sesion3_4.objects.all().order_by('?').first()),
                  ('Campo de texto 3046',Nivel4_Sesion3_5.objects.all().order_by('?').first()),
                  ('Campo de texto 3047',Nivel4_Sesion3_6.objects.all().order_by('?').first())
        ]
        fdf = forge_fdf("",fields,[],[],[])
        fdf_file = open("C:\\Temp\\datas3_n4.fdf","wb")
        fdf_file.write(fdf)
        fdf_file.close()
        os.system('pdftk C:\\Temp\\sesion3_n4_plantilla.pdf fill_form C:\\Temp\\datas3_n4.fdf output C:\\Temp\\sesion3_4.pdf')
        nuevo.soporte = File(open("C://Temp//sesion3_4.pdf", 'rb'))
        nuevo.save()
        evidencia_docente.soporte = nuevo
        evidencia_docente.save()
generar_nivel4_3.short_description = "Generar Sesion 3 - Nivel 4"

class Nivel4_Sesion3_1Admin(admin.ModelAdmin):
    list_display = ['respuesta']
    actions = [generar_nivel4_3]
admin.site.register(Nivel4_Sesion3_1,Nivel4_Sesion3_1Admin)







def generar_nivel3_3(modeladmin,request,queryset):
    evidencia_docentes = EvidenciaDocentes.objects.filter(entregable__id=40,soporte=None)
    for evidencia_docente in evidencia_docentes:
        nuevo = SoporteEntregableDocente(grupo=evidencia_docente.participante.grupo,entregable=EntregableDocentes.objects.get(id=40))
        fields = [('Campo de texto 3033',Nivel3_Sesion3_1.objects.all().order_by('?').first().respuesta),
                  ('Campo de texto 3034',Nivel3_Sesion3_2.objects.all().order_by('?').first().respuesta)
        ]
        fdf = forge_fdf("",fields,[],[],[])
        fdf_file = open("C:\\Temp\\datas3_n3.fdf","wb")
        fdf_file.write(fdf)
        fdf_file.close()
        os.system('pdftk C:\\Temp\\sesion3_n3_plantilla.pdf fill_form C:\\Temp\\datas3_n3.fdf output C:\\Temp\\sesion3_3.pdf')
        nuevo.soporte = File(open("C://Temp//sesion3_3.pdf", 'rb'))
        nuevo.save()
        evidencia_docente.soporte = nuevo
        evidencia_docente.save()
generar_nivel3_3.short_description = "Generar Sesion 3 - Nivel 3"

class Nivel3_Sesion3_1Admin(admin.ModelAdmin):
    list_display = ['respuesta']
    actions = [generar_nivel3_3]
admin.site.register(Nivel3_Sesion3_1,Nivel3_Sesion3_1Admin)
admin.site.register(Nivel3_Sesion3_2)









class Nivel4_Sesion3_2Admin(admin.ModelAdmin):
    list_display = ['respuesta']
    actions = []
admin.site.register(Nivel4_Sesion3_2,Nivel4_Sesion3_2Admin)

class Nivel4_Sesion3_3Admin(admin.ModelAdmin):
    list_display = ['respuesta']
    actions = []
admin.site.register(Nivel4_Sesion3_3,Nivel4_Sesion3_3Admin)

class Nivel4_Sesion3_4Admin(admin.ModelAdmin):
    list_display = ['respuesta']
    actions = [generar_nivel4_2]
admin.site.register(Nivel4_Sesion3_4,Nivel4_Sesion3_4Admin)

class Nivel4_Sesion3_5Admin(admin.ModelAdmin):
    list_display = ['respuesta']
    actions = []
admin.site.register(Nivel4_Sesion3_5,Nivel4_Sesion3_5Admin)

class Nivel4_Sesion3_6Admin(admin.ModelAdmin):
    list_display = ['respuesta']
    actions = []
admin.site.register(Nivel4_Sesion3_6,Nivel4_Sesion3_6Admin)





admin.site.register(Nivel1_Sesion1_1,Nivel1_Sesion1_1Admin)
admin.site.register(Nivel1_Sesion1_2)
admin.site.register(Nivel1_Sesion1_3)
admin.site.register(Nivel1_Sesion1_4)
admin.site.register(Nivel1_Sesion1_5)
admin.site.register(Nivel1_Sesion1_6)
admin.site.register(Nivel1_Sesion1_7)
admin.site.register(Nivel1_Sesion1_8)
admin.site.register(Nivel1_Sesion1_9)
admin.site.register(Nivel1_Sesion1_10)
admin.site.register(Nivel1_Sesion1_11)
admin.site.register(Nivel1_Sesion1_12)
admin.site.register(Nivel1_Sesion1_REDA)
admin.site.register(Nivel1_Sesion2_1)
admin.site.register(Nivel1_Sesion2_2)
admin.site.register(Nivel1_Sesion2_3)
admin.site.register(Nivel1_Sesion2_4)

def generar_nivel1_sesion4(modeladmin,request,queryset):
    evidencia_docentes = EvidenciaDocentes.objects.filter(entregable__id=17,soporte=None)
    for evidencia_docente in evidencia_docentes:
        nuevo = SoporteEntregableDocente(grupo=evidencia_docente.participante.grupo,entregable=EntregableDocentes.objects.get(id=17))
        pregunta1 = Nivel1_Sesion4_preguntas_aleatorias.objects.order_by('?')[0]
        pregunta2 = Nivel1_Sesion4_preguntas_aleatorias.objects.order_by('?')[0]
        pregunta3 = Nivel1_Sesion4_preguntas_aleatorias.objects.order_by('?')[0]
        pregunta4 = Nivel1_Sesion4_preguntas_aleatorias.objects.order_by('?')[0]
        pregunta5 = Nivel1_Sesion4_preguntas_aleatorias.objects.order_by('?')[0]
        pregunta6 = Nivel1_Sesion4_preguntas_aleatorias.objects.order_by('?')[0]
        fields = [('Campo de texto 165',evidencia_docente.participante.nombres+ " "+evidencia_docente.participante.apellidos),
                  ('Campo de texto 166',evidencia_docente.participante.radicado.nombre_ie),
                  ('Campo de texto 167',evidencia_docente.participante.radicado.nombre_sede),
                  ('Campo de texto 168',evidencia_docente.participante.grupo.municipio.nombre),

                  ('Campo de texto 169',pregunta1.pregunta),
                  ('Campo de texto 175',pregunta1.respuesta),
                  ('Campo de texto 176',pregunta1.incorrecta_1),
                  ('Campo de texto 177',pregunta1.incorrecta_2),

                  ('Campo de texto 170',pregunta2.pregunta),
                  ('Campo de texto 178',pregunta2.respuesta),
                  ('Campo de texto 179',pregunta2.incorrecta_1),
                  ('Campo de texto 180',pregunta2.incorrecta_2),

                  ('Campo de texto 171',pregunta3.pregunta),
                  ('Campo de texto 181',pregunta3.respuesta),
                  ('Campo de texto 182',pregunta3.incorrecta_1),
                  ('Campo de texto 183',pregunta3.incorrecta_2),

                  ('Campo de texto 172',pregunta4.pregunta),
                  ('Campo de texto 184',pregunta4.respuesta),
                  ('Campo de texto 185',pregunta4.incorrecta_1),
                  ('Campo de texto 186',pregunta4.incorrecta_2),

                  ('Campo de texto 173',pregunta5.pregunta),
                  ('Campo de texto 187',pregunta5.respuesta),
                  ('Campo de texto 188',pregunta5.incorrecta_1),
                  ('Campo de texto 189',pregunta5.incorrecta_2),

                  ('Campo de texto 174',pregunta6.pregunta),
                  ('Campo de texto 190',pregunta6.respuesta),
                  ('Campo de texto 191',pregunta6.incorrecta_1),
                  ('Campo de texto 192',pregunta6.incorrecta_2),

                  ('Campo de texto 194',Nivel1_Sesion4_1.objects.order_by('?')[0].respuesta),
                  ('Campo de texto 197',Nivel1_Sesion4_2.objects.order_by('?')[0].respuesta),
                  ('Campo de texto 198',Nivel1_Sesion4_3.objects.order_by('?')[0].respuesta),
                  ('Campo de texto 199',Nivel1_Sesion4_4.objects.order_by('?')[0].respuesta),
                  ('Campo de texto 200',Nivel1_Sesion4_5.objects.order_by('?')[0].respuesta),
        ]
        fdf = forge_fdf("",fields,[],[],[])
        fdf_file = open("C:\\Temp\\datas4.fdf","wb")
        fdf_file.write(fdf)
        fdf_file.close()
        os.system('pdftk C:\\Temp\\sesion4_n1_plantilla.pdf fill_form C:\\Temp\\datas4.fdf output C:\\Temp\\sesion4_1.pdf flatten')
        nuevo.soporte = File(open("C://Temp//sesion4_1.pdf", 'rb'))
        nuevo.save()
        evidencia_docente.soporte = nuevo
        evidencia_docente.save()
generar_nivel1_sesion4.short_description = "Generar Sesion 4 - Nivel 1"

class Nivel1_Sesion4_preguntas_aleatoriasAdmin(admin.ModelAdmin):
    list_display = ['respuesta']
    actions = [generar_nivel1_sesion4]

admin.site.register(Nivel1_Sesion4_preguntas_aleatorias,Nivel1_Sesion4_preguntas_aleatoriasAdmin)





def generar_nivel3_sesion1(modeladmin,request,queryset):
    evidencia_docentes = EvidenciaDocentes.objects.filter(entregable__id=35,soporte=None)
    for evidencia_docente in evidencia_docentes:
        nuevo = SoporteEntregableDocente(grupo=evidencia_docente.participante.grupo,entregable=EntregableDocentes.objects.get(id=35))
        fields = [('Campo de texto 3014',Nivel3_Sesion1_1.objects.order_by('?').first().respuesta),
        ]
        fdf = forge_fdf("",fields,[],[],[])
        fdf_file = open("C:\\Temp\\datas1_1_n3.fdf","wb")
        fdf_file.write(fdf)
        fdf_file.close()
        os.system('pdftk C:\\Temp\\sesion1_1_n3_plantilla.pdf fill_form C:\\Temp\\datas1_1_n3.fdf output C:\\Temp\\sesion1_1_3.pdf')
        nuevo.soporte = File(open("C://Temp//sesion1_1_3.pdf", 'rb'))
        nuevo.save()
        evidencia_docente.soporte = nuevo
        evidencia_docente.save()
generar_nivel3_sesion1.short_description = "Generar Sesion 1.1 - Nivel 3"

class Nivel3_Sesion1_1Admin(admin.ModelAdmin):
    list_display = ['respuesta']
    actions = [generar_nivel3_sesion1]

admin.site.register(Nivel3_Sesion1_1,Nivel3_Sesion1_1Admin)




def generar_nivel3_sesion1_2(modeladmin,request,queryset):
    evidencia_docentes = EvidenciaDocentes.objects.filter(entregable__id=36,soporte=None)
    for evidencia_docente in evidencia_docentes:
        nuevo = SoporteEntregableDocente(grupo=evidencia_docente.participante.grupo,entregable=EntregableDocentes.objects.get(id=36))
        fields = [('Campo de texto 3015',Nivel3_Sesion1_2.objects.order_by('?').first().respuesta),
        ]
        fdf = forge_fdf("",fields,[],[],[])
        fdf_file = open("C:\\Temp\\datas1_2_n3.fdf","wb")
        fdf_file.write(fdf)
        fdf_file.close()
        os.system('pdftk C:\\Temp\\sesion1_2_n3_plantilla.pdf fill_form C:\\Temp\\datas1_2_n3.fdf output C:\\Temp\\sesion1_2_3.pdf')
        nuevo.soporte = File(open("C://Temp//sesion1_2_3.pdf", 'rb'))
        nuevo.save()
        evidencia_docente.soporte = nuevo
        evidencia_docente.save()
generar_nivel3_sesion1_2.short_description = "Generar Sesion 1.2 - Nivel 3"


def certificaciones(modeladmin,request,queryset):
    from PIL import Image
    from PIL import ImageFont
    from PIL import ImageDraw

    for participante in ParticipanteEscuelaTic.objects.all():
        img = Image.open("C:\\Temp\\Diploma.png")
        draw = ImageDraw.Draw(img)
        font = ImageFont.truetype("arial.ttf", 50)
        draw.text((410, 230),participante.nombres+" "+participante.apellidos,(255,255,255),font=font)

        font = ImageFont.truetype("arial.ttf", 18)
        draw.text((762, 360),str(participante.cedula),(29,42,64),font=font)

        img.save('C:\\Diploma\\'+str(participante.cedula)+'.png')

certificaciones.short_description = "Generar certificacion Escuela Tic"



class Nivel3_Sesion1_2Admin(admin.ModelAdmin):
    list_display = ['respuesta']
    actions = [generar_nivel3_sesion1_2,certificaciones]

admin.site.register(Nivel3_Sesion1_2,Nivel3_Sesion1_2Admin)




'''
def generar_nivel3_sesion2_1(modeladmin,request,queryset):
    evidencia_docentes = EvidenciaDocentes.objects.filter(entregable__id=37,soporte=None)
    for evidencia_docente in evidencia_docentes:
        nuevo = SoporteEntregableDocente(grupo=evidencia_docente.participante.grupo,entregable=EntregableDocentes.objects.get(id=37))

        document = Document()
        document.add_paragraph(unicode(evidencia_docente.participante.nombres).lower() + " " + unicode(evidencia_docente.participante.apellidos).lower())
        conclusiones = Nivel3_Sesion2_1.objects.order_by('?')[:3]

        for conclusion in conclusiones:
            document.add_paragraph(unicode(conclusion.respuesta))

        document.save("C://Temp//sesion2_1_3.docx")

        nuevo.soporte = File(open("C://Temp//sesion2_1_3.docx", 'rb'))
        nuevo.save()
        evidencia_docente.soporte = nuevo
        evidencia_docente.save()
generar_nivel3_sesion2_1.short_description = "Generar Sesion 2.1 - Nivel 3"
'''
class Nivel3_Sesion2_1Admin(admin.ModelAdmin):
    list_display = ['respuesta']
    actions = []

admin.site.register(Nivel3_Sesion2_1,Nivel3_Sesion2_1Admin)


def carga_masiva_docentes_n3_n4(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Carga Masiva N3 y N4.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        #logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        #logo.drawing.top = 10
        #logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Carga Masiva N3 y N4"
        #hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Carga Masiva N3 y N4'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['SOPORTE',30]),
                   tuple(['RESULTADO',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]





        hoja2 = archivo.get_sheet_by_name('hoja2')
        hoja2.title = "RED"
        #hoja1.add_image(logo)

        celda = hoja2.cell('E2')
        celda.value = 'Formacion'

        celda = hoja2.cell('E3')
        celda.value = 'RED'

        celda = hoja2.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja2.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num_2 = 5

        columns = [tuple(['SOPORTE',30]),
                   tuple(['NOMBRE',30]),
                   tuple(['CEDULA',30]),
                   tuple(['GRUPO',30]),
                   tuple(['RESULTADO',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja2.cell(row=row_num_2, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja2.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]



        soportes = ZipFile(settings.MEDIA_ROOT+'//'+str(archivo_queryset.archivo),'r')

        i = 0

        for soporte in soportes.namelist():
            resultado_soporte = ""
            cedula = str(soporte).split('.')[0]

            try:
                participante = ParticipanteDocenteMasivo.objects.get(cedula=cedula)
            except:
                resultado_soporte = "No existe cedula Base generar"
            else:
                resultado_soporte = "Grupo identificado"
                participantes = ParticipanteDocenteMasivo.objects.filter(codigo_masivo=participante.codigo_masivo,subgrupo=participante.subgrupo)

                source = soportes.open(soporte)
                target = file(os.path.join(r"C:\Temp\masivo",unicode(cedula)+".pdf"),"wb")
                with source, target:
                    shutil.copyfileobj(source,target)

                output1 = PdfFileWriter()
                output2 = PdfFileWriter()
                output3 = PdfFileWriter()
                output4 = PdfFileWriter()
                output5 = PdfFileWriter()
                output6 = PdfFileWriter()
                output7 = PdfFileWriter()

                with open("C:\\Temp\\masivo\\"+unicode(cedula)+".pdf", "rb") as f:
                    input = PdfFileReader(f,'rb')
                    if participantes.count() <= 11:
                        output1.addPage(input.getPage(0))
                        output2.addPage(input.getPage(1))
                        output3.addPage(input.getPage(2))
                        output4.addPage(input.getPage(3))
                        output5.addPage(input.getPage(4))
                        output6.addPage(input.getPage(5))
                        output7.addPage(input.getPage(6))
                    else:
                        output1.addPage(input.getPage(0))
                        output1.addPage(input.getPage(1))
                        output2.addPage(input.getPage(2))
                        output2.addPage(input.getPage(3))
                        output3.addPage(input.getPage(4))
                        output3.addPage(input.getPage(5))
                        output4.addPage(input.getPage(6))
                        output4.addPage(input.getPage(7))
                        output5.addPage(input.getPage(8))
                        output5.addPage(input.getPage(9))
                        output6.addPage(input.getPage(10))
                        output6.addPage(input.getPage(11))
                        output7.addPage(input.getPage(12))
                        output7.addPage(input.getPage(13))

                    with open("C:\\Temp\\masivo\\"+unicode(participante.cedula)+"_3_1.pdf", "wb") as outputStream1:
                        output1.write(outputStream1)

                    with open("C:\\Temp\\masivo\\"+unicode(participante.cedula)+"_3_2.pdf", "wb") as outputStream2:
                        output2.write(outputStream2)

                    with open("C:\\Temp\\masivo\\"+unicode(participante.cedula)+"_3_3.pdf", "wb") as outputStream3:
                        output3.write(outputStream3)

                    with open("C:\\Temp\\masivo\\"+unicode(participante.cedula)+"_4_1.pdf", "wb") as outputStream4:
                        output4.write(outputStream4)

                    with open("C:\\Temp\\masivo\\"+unicode(participante.cedula)+"_4_2.pdf", "wb") as outputStream5:
                        output5.write(outputStream5)

                    with open("C:\\Temp\\masivo\\"+unicode(participante.cedula)+"_4_3.pdf", "wb") as outputStream6:
                        output6.write(outputStream6)

                    with open("C:\\Temp\\masivo\\"+unicode(participante.cedula)+"_4_4.pdf", "wb") as outputStream7:
                        output7.write(outputStream7)




                codigo = CodigoMasivo_Docentes.objects.get(codigo = participante.codigo_masivo.codigo)
                formador = Formador.objects.get(cedula=codigo.cedula)

                try:
                    grupo = GrupoDocentes.objects.filter(formador__cedula = formador.cedula).get(nombre = participante.grupo)
                except:
                    grupo = GrupoDocentes.objects.filter(formador__cedula = formador.cedula)[0]


                nuevo_soporte_3_1 = SoporteEntregableDocente(grupo=grupo,
                                                                 entregable=EntregableDocentes.objects.get(id=29),
                                                                 soporte=File(open("C://Temp//masivo//" + participante.cedula+"_3_1.pdf", 'rb')))
                nuevo_soporte_3_1.save()

                nuevo_soporte_3_2 = SoporteEntregableDocente(grupo=grupo,
                                                                 entregable=EntregableDocentes.objects.get(id=31),
                                                                 soporte=File(open("C://Temp//masivo//" + participante.cedula+"_3_2.pdf", 'rb')))
                nuevo_soporte_3_2.save()

                nuevo_soporte_3_3 = SoporteEntregableDocente(grupo=grupo,
                                                                 entregable=EntregableDocentes.objects.get(id=33),
                                                                 soporte=File(open("C://Temp//masivo//" + participante.cedula+"_3_3.pdf", 'rb')))
                nuevo_soporte_3_3.save()



                nuevo_soporte_4_1 = SoporteEntregableDocente(grupo=grupo,
                                                                 entregable=EntregableDocentes.objects.get(id=41),
                                                                 soporte=File(open("C://Temp//masivo//" + participante.cedula+"_4_1.pdf", 'rb')))
                nuevo_soporte_4_1.save()

                nuevo_soporte_4_2 = SoporteEntregableDocente(grupo=grupo,
                                                                 entregable=EntregableDocentes.objects.get(id=43),
                                                                 soporte=File(open("C://Temp//masivo//" + participante.cedula+"_4_2.pdf", 'rb')))
                nuevo_soporte_4_2.save()

                nuevo_soporte_4_3 = SoporteEntregableDocente(grupo=grupo,
                                                                 entregable=EntregableDocentes.objects.get(id=45),
                                                                 soporte=File(open("C://Temp//masivo//" + participante.cedula+"_4_3.pdf", 'rb')))
                nuevo_soporte_4_3.save()

                nuevo_soporte_4_5 = SoporteEntregableDocente(grupo=grupo,
                                                                 entregable=EntregableDocentes.objects.get(id=49),
                                                                 soporte=File(open("C://Temp//masivo//" + participante.cedula+"_4_4.pdf", 'rb')))
                nuevo_soporte_4_5.save()

                for parti in participantes:
                    row_num_2 += 1

                    try:
                        participante_principal = ParticipanteDocente.objects.get(cedula=parti.cedula)
                    except:
                        resultado = "No existe la cedula en la base principal"
                    else:
                        resultado = "Cargado"
                        participante_principal.formador = formador
                        participante_principal.grupo = grupo
                        participante_principal.save()

                        evidencia1 = EvidenciaDocentes.objects.filter(participante__cedula = parti.cedula).get(entregable__id=29)
                        evidencia1.soporte = nuevo_soporte_3_1
                        evidencia1.save()

                        evidencia2 = EvidenciaDocentes.objects.filter(participante__cedula = parti.cedula).get(entregable__id=31)
                        evidencia2.soporte = nuevo_soporte_3_2
                        evidencia2.save()

                        evidencia3 = EvidenciaDocentes.objects.filter(participante__cedula = parti.cedula).get(entregable__id=33)
                        evidencia3.soporte = nuevo_soporte_3_3
                        evidencia3.save()


                        evidencia4 = EvidenciaDocentes.objects.filter(participante__cedula = parti.cedula).get(entregable__id=41)
                        evidencia4.soporte = nuevo_soporte_4_1
                        evidencia4.save()

                        evidencia5 = EvidenciaDocentes.objects.filter(participante__cedula = parti.cedula).get(entregable__id=43)
                        evidencia5.soporte = nuevo_soporte_4_2
                        evidencia5.save()

                        evidencia6 = EvidenciaDocentes.objects.filter(participante__cedula = parti.cedula).get(entregable__id=45)
                        evidencia6.soporte = nuevo_soporte_4_3
                        evidencia6.save()

                        evidencia7 = EvidenciaDocentes.objects.filter(participante__cedula = parti.cedula).get(entregable__id=49)
                        evidencia7.soporte = nuevo_soporte_4_5
                        evidencia7.save()




                    row = [
                        soporte,
                        parti.nombre,
                        parti.cedula,
                        grupo.nombre,
                        resultado
                    ]

                    for col_num in xrange(len(row)):
                        c = hoja2.cell(row=row_num_2, column=col_num+1)
                        if row[col_num] == True:
                            c.value = "SI"
                        if row[col_num] == False:
                            c.value = "NO"
                        if row[col_num] == None:
                            c.value = ""
                        else:
                            c.value = row[col_num]
                        c.style = co



            row_num += 1
            row = [
                soporte,
                resultado_soporte
            ]

            for col_num in xrange(len(row)):
                c = hoja1.cell(row=row_num, column=col_num+1)
                if row[col_num] == True:
                    c.value = "SI"
                if row[col_num] == False:
                    c.value = "NO"
                if row[col_num] == None:
                    c.value = ""
                else:
                    c.value = row[col_num]
                c.style = co


        archivo.save(response)
        return response
carga_masiva_docentes_n3_n4.short_description = "Cargar masiva docentes n3 y n4"


class CargaMasiva_n3_n4Admin(admin.ModelAdmin):
    list_display = ['id','archivo']
    actions = [carga_masiva_docentes_n3_n4]

admin.site.register(CargaMasiva_n3_n4,CargaMasiva_n3_n4Admin)