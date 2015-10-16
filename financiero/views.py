#!/usr/bin/env python
# -*- coding: utf-8 -*-
from django.views.generic import TemplateView, CreateView, UpdateView
from django.views.generic.edit import ModelFormMixin
from mixins.mixins import FinancieroMixin
from region.models import Region
from gestor.models import Gestor
from gestor.forms import NuevoForm
from formador.models import Formador
from formador.forms import NuevoForm as NuevoFormFormador
from funcionario.forms import NuevoFuncionarioForm
from funcionario.models import Funcionario
from acceso.models import Corte
from acceso.forms import CorteForm
from .forms import LiquidacionGestorForm, LiquidacionFormadorForm
from acceso.models import Evidencia,Actividad
from django.db.models import Sum
from django.core.mail import EmailMessage
import StringIO
from django_datatables_view.base_datatable_view import BaseDatatableView
from django.db.models import Q
from .models import LiquidacionGestor, LiquidacionFormador
from money import Money
import datetime

import pythoncom
from win32com import client
from win32com.client import DispatchEx
from win32com import *
import os
import zipfile
import shutil

from django.http import HttpResponse
import time
from conf import settings
import openpyxl
from django.utils.encoding import smart_unicode

from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font

t = Style(font=Font(name='Calibri',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='C9C9C9',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

co = Style(font=Font(name='Calibri',size=11),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

co_money = Style(font=Font(name='Calibri',size=11),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='Money')

v = Style(font=Font(name='Calibri',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='E4F5E1',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

vc = Style(font=Font(name='Calibri',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='D4F5CE',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

class FinancieroView(FinancieroMixin,TemplateView):
    template_name = 'financiero.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        return super(FinancieroView,self).get_context_data(**kwargs)

class GestorView(FinancieroMixin,TemplateView):
    template_name = 'listado_gestores_financiero.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        return super(GestorView,self).get_context_data(**kwargs)

class GestorCorteEvidenciaView(FinancieroMixin,TemplateView):
    template_name = 'listado_gestores_evidencia_corte.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['ID_CORTE'] = self.kwargs['corte_id']
        kwargs['ID_GESTOR'] = self.kwargs['gestor_id']
        kwargs['CORTE'] = Corte.objects.get(pk=self.kwargs['corte_id']).titulo
        kwargs['GESTOR'] = Gestor.objects.get(pk=self.kwargs['gestor_id']).nombre
        return super(GestorCorteEvidenciaView,self).get_context_data(**kwargs)

class FormadorView(FinancieroMixin,TemplateView):
    template_name = 'tipo_formador_financiero.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        return super(FormadorView,self).get_context_data(**kwargs)

class FormadorTipoView(FinancieroMixin,TemplateView):
    template_name = 'listado_formadores_financiero.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['ID_TIPO'] = self.kwargs['tipo_id']
        return super(FormadorTipoView,self).get_context_data(**kwargs)

class FuncionarioView(FinancieroMixin,TemplateView):
    template_name = 'listado_funcionario_financiero.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        return super(FuncionarioView,self).get_context_data(**kwargs)

class NuevoGestorView(FinancieroMixin, CreateView):
    model = Gestor
    form_class = NuevoForm
    success_url = "../"
    template_name = "nuevo_gestor.html"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        return super(NuevoGestorView,self).get_context_data(**kwargs)

class NuevoFormadorView(FinancieroMixin, CreateView):
    model = Formador
    form_class = NuevoFormFormador
    success_url = "../"
    template_name = "nuevo_formador.html"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['ID_TIPO'] = self.kwargs['tipo_id']
        return super(NuevoFormadorView,self).get_context_data(**kwargs)

class NuevoFuncionarioView(FinancieroMixin, CreateView):
    model = Funcionario
    form_class = NuevoFuncionarioForm
    success_url = "../"
    template_name = "nuevo_funcionario.html"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        return super(NuevoFuncionarioView,self).get_context_data(**kwargs)

class NuevoCorteView(FinancieroMixin, CreateView):
    model = Corte
    form_class = CorteForm
    success_url = "../"
    template_name = "nuevo_corte.html"

    def get_context_data(self, **kwargs):
        reporte = Evidencia.objects.filter(radicado__region__id=self.kwargs['pk']).filter(corte=None).exclude(soporte="")
        kwargs['CANTIDAD'] = reporte.count()
        kwargs['VALOR'] = "$ "+str(reporte.aggregate(Sum('valor__valor'))['valor__valor__sum'])
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']

        return super(NuevoCorteView,self).get_context_data(**kwargs)

    def form_valid(self, form):
        self.object = form.save()
        corte = self.object
        reporte = Evidencia.objects.filter(radicado__region__id=self.kwargs['pk']).filter(corte=None).exclude(soporte="")
        for evidencia in reporte:
            evidencia.corte = corte
            evidencia.save()
        return super(ModelFormMixin, self).form_valid(form)

def reporte_quincenal_financiero(request,pk):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte de Quincenas.xlsx'
    archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

    logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
    logo.drawing.top = 10
    logo.drawing.left = 25

    hoja1 = archivo.get_sheet_by_name('hoja1')
    hoja1.title = "Informe Quincenas Gestores"
    hoja1.add_image(logo)

    celda = hoja1.cell('E2')
    celda.value = 'ADMINISTRATIVO Y FINANCIERO'

    celda = hoja1.cell('E3')
    celda.value = 'REPORTE QUINCENAS GESTORES'

    celda = hoja1.cell('I3')
    celda.value = time.strftime("%d/%m/%y")

    celda = hoja1.cell('I4')
    celda.value = time.strftime("%I:%M:%S %p")

    row_num = 5

    cortes = Corte.objects.filter(region__id=pk)

    columns = [tuple(['Nombre Gestor',60]),
               tuple(['Banco',30]),
               tuple(['Tipo de Cuenta',30]),
               tuple(['Numero de Cuenta',30])
               ]
    for corte in cortes:
        columns.append(tuple([corte.titulo+" - "+str(corte.fecha),90]))



    for col_num in xrange(len(columns)):
        c = hoja1.cell(row=row_num, column=col_num+1)
        c.value = columns[col_num][0]
        c.style = t
        hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


    gestores = Gestor.objects.filter(region__id=pk)

    for gestor in gestores:
        row_num += 1
        row = [
            gestor.nombre,
            gestor.banco,
            gestor.tipo_cuenta,
            gestor.numero_cuenta,
        ]

        for corte in cortes:
            row.append(Evidencia.objects.filter(gestor__id=gestor.id).exclude(corte=None).filter(corte=corte).aggregate(Sum('valor__valor'))['valor__valor__sum'])

        for col_num in xrange(len(row)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            if row[col_num] == True:
                c.value = "SI"
            if row[col_num] == False:
                c.value = "NO"
            if row[col_num] == None:
                c.value = ""
            else:
                c.value = row[col_num]
            c.style = co

        c = hoja1.cell(row=6, column=6)
        c.style = co_money

    archivo.save(response)
    return response

def reporte_gestor(request,pk,corte_id,gestor_id):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    gestor = Gestor.objects.get(id=gestor_id)
    corte = Corte.objects.get(id=corte_id)
    response['Content-Disposition'] = 'attachment; filename='+smart_unicode(gestor.nombre)+'.xlsx'
    archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

    logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
    logo.drawing.top = 10
    logo.drawing.left = 25

    hoja1 = archivo.get_sheet_by_name('hoja1')
    hoja1.title = smart_unicode(gestor.nombre)
    hoja1.add_image(logo)

    celda = hoja1.cell('E2')
    celda.value = 'ACCESO'

    celda = hoja1.cell('E3')
    celda.value = 'REPORTE QUINCENAL - '+smart_unicode(gestor.nombre)

    celda = hoja1.cell('I3')
    celda.value = time.strftime("%d/%m/%y")

    celda = hoja1.cell('I4')
    celda.value = time.strftime("%I:%M:%S %p")

    row_num = 5

    columnas = Actividad.objects.order_by('id').values('nombre','titulo')
    columns = [tuple(['Radicado',30]),
               tuple(['Departamento',30]),
               tuple(['Municipio',30]),
               tuple(['Ciclo',30]),
               tuple(['Componente',30]),
               tuple(['Modulo',30]),
               tuple(['Actividad',30]),
               tuple(['Valor',30]),
               tuple(['Corte',30]),
               ]



    for col_num in xrange(len(columns)):
        c = hoja1.cell(row=row_num, column=col_num+1)
        c.value = columns[col_num][0]
        c.style = t
        hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


    evidencias = Evidencia.objects.filter(gestor__id=gestor_id).filter(corte_id=corte.id)

    for evidencia in evidencias:
        row_num += 1
        row = [
            evidencia.radicado.numero,
            evidencia.radicado.municipio.departamento.nombre,
            evidencia.radicado.municipio.nombre,
            evidencia.ciclo.nombre,
            evidencia.componente.nombre,
            evidencia.modulo.nombre,
            str(evidencia.actividad.nombre)+" - "+str(evidencia.actividad.titulo),
            evidencia.valor.valor,
            str(evidencia.corte.fecha)+" - "+str(evidencia.corte.titulo) if evidencia.corte != None else "",
        ]

        for col_num in xrange(len(row)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            if row[col_num] == True:
                c.value = "SI"
            if row[col_num] == False:
                c.value = "NO"
            if row[col_num] == None:
                c.value = ""
            else:
                c.value = row[col_num]
            c.style = co



    archivo.save(response)
    return response

def reporte_gestor_email(request,pk,corte_id,gestor_id):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    gestor = Gestor.objects.get(id=gestor_id)
    corte = Corte.objects.get(id=corte_id)
    response['Content-Disposition'] = 'attachment; filename='+smart_unicode(gestor.nombre)+'.xlsx'
    archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

    logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
    logo.drawing.top = 10
    logo.drawing.left = 25

    hoja1 = archivo.get_sheet_by_name('hoja1')
    hoja1.title = smart_unicode(gestor.nombre)
    hoja1.add_image(logo)

    celda = hoja1.cell('E2')
    celda.value = 'ACCESO'

    celda = hoja1.cell('E3')
    celda.value = 'REPORTE QUINCENAL - '+smart_unicode(gestor.nombre)

    celda = hoja1.cell('I3')
    celda.value = time.strftime("%d/%m/%y")

    celda = hoja1.cell('I4')
    celda.value = time.strftime("%I:%M:%S %p")

    row_num = 5

    columnas = Actividad.objects.order_by('id').values('nombre','titulo')
    columns = [tuple(['Radicado',30]),
               tuple(['Departamento',30]),
               tuple(['Municipio',30]),
               tuple(['Ciclo',30]),
               tuple(['Componente',30]),
               tuple(['Modulo',30]),
               tuple(['Actividad',30]),
               tuple(['Valor',30]),
               tuple(['Corte',30]),
               ]



    for col_num in xrange(len(columns)):
        c = hoja1.cell(row=row_num, column=col_num+1)
        c.value = columns[col_num][0]
        c.style = t
        hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


    evidencias = Evidencia.objects.filter(gestor__id=gestor_id).filter(corte_id=corte.id)
    valor = evidencias.aggregate(Sum('valor__valor'))['valor__valor__sum']

    for evidencia in evidencias:
        row_num += 1
        row = [
            evidencia.radicado.numero,
            evidencia.radicado.municipio.departamento.nombre,
            evidencia.radicado.municipio.nombre,
            evidencia.ciclo.nombre,
            evidencia.componente.nombre,
            evidencia.modulo.nombre,
            str(evidencia.actividad.nombre)+" - "+str(evidencia.actividad.titulo),
            evidencia.valor.valor,
            str(evidencia.corte.fecha)+" - "+str(evidencia.corte.titulo) if evidencia.corte != None else "",
        ]

        for col_num in xrange(len(row)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            if row[col_num] == True:
                c.value = "SI"
            if row[col_num] == False:
                c.value = "NO"
            if row[col_num] == None:
                c.value = ""
            else:
                c.value = row[col_num]
            c.style = co

    f = StringIO.StringIO()
    archivo.save(f)

    contenido = '<b>Apreciado(a) Gestor(a): '+gestor.nombre+'</b><br><br>'+'<p>Cordial saludo,</p>'+'<br>'+\
                '<p>Para la <b>Asociación Nacional para el Desarrollo Social - ANDES</b> y el programa <b>Computadores para Educar</b> es ' \
                'un placer contar con Gestores tan comprometidos con su trabajo, te informamos que adjunto puedes encontrar el reporte detallado con las' \
                ' actividades cargadas en el sistema SICAN con corte: <b>'+str(corte.fecha)+'</b> por un valor total de' \
                ' <b>$'+str(int(valor))+'</b>.</p>'+'<br>'
    email = EmailMessage("Reporte Quincena - "+corte.titulo+" - "+str(corte.fecha),contenido,to=[gestor.correo])
    email.attach('Reporte.xlsx', f.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    email.content_subtype = "html"
    email.send()

    archivo.save(response)
    return response

class DocumentalView(FinancieroMixin,TemplateView):
    template_name = 'documental.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        return super(DocumentalView,self).get_context_data(**kwargs)

class LiquidacionesView(FinancieroMixin,TemplateView):
    template_name = 'financiero_liquidaciones.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        return super(LiquidacionesView,self).get_context_data(**kwargs)

class LiquidacionAccesoView(FinancieroMixin,TemplateView):
    template_name = 'financiero_liquidaciones_acceso.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        return super(LiquidacionAccesoView,self).get_context_data(**kwargs)

class LiquidacionFormacionView(FinancieroMixin,TemplateView):
    template_name = 'financiero_liquidaciones_formacion.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        return super(LiquidacionFormacionView,self).get_context_data(**kwargs)

class LiquidacionAccesoTableView(BaseDatatableView):
    model = LiquidacionGestor
    columns = [
        'id',
        'gestor',
        'fecha_terminacion',
        'contrato',
        'valor_inicial',
        'valor_ejecutado',
        'valor_pagado'
    ]

    order_columns = [
        'id',
        'gestor',
        'fecha_terminacion',
        'contrato',
        'valor_inicial',
        'valor_ejecutado',
        'valor_pagado'
    ]

    def get_initial_queryset(self):
        if not self.model:
            raise NotImplementedError("Need to provide a model or implement get_initial_queryset!")
        return self.model.objects.filter(gestor__region__id=self.kwargs['pk'])

    def filter_queryset(self, qs):
        search = self.request.GET.get(u'search[value]', None)
        q = Q()
        if search:
            q |= Q(**{'gestor__nombre__icontains' : search})
            qs = qs.filter(q)
        return qs

    def render_column(self, row, column):
        if column == 'gestor':
            return row.gestor.nombre
        if column == 'fecha_terminacion':
            return row.fecha_terminacion.strftime('%d/%m/%Y')
        else:
            return super(LiquidacionAccesoTableView,self).render_column(row,column)

class LiquidacionFormacionTableView(BaseDatatableView):
    model = LiquidacionFormador
    columns = [
        'id',
        'formador',
        'fecha_terminacion',
        'contrato',
        'valor_inicial',
        'valor_ejecutado',
        'valor_pagado'
    ]

    order_columns = [
        'id',
        'formador',
        'fecha_terminacion',
        'contrato',
        'valor_inicial',
        'valor_ejecutado',
        'valor_pagado'
    ]

    def get_initial_queryset(self):
        if not self.model:
            raise NotImplementedError("Need to provide a model or implement get_initial_queryset!")
        return self.model.objects.filter(formador__region__id=self.kwargs['pk'])

    def filter_queryset(self, qs):
        search = self.request.GET.get(u'search[value]', None)
        q = Q()
        if search:
            q |= Q(**{'formador__nombre__icontains' : search})
            qs = qs.filter(q)
        return qs

    def render_column(self, row, column):
        if column == 'formador':
            return row.formador.nombre
        if column == 'fecha_terminacion':
            return row.fecha_terminacion.strftime('%d/%m/%Y')
        else:
            return super(LiquidacionFormacionTableView,self).render_column(row,column)

class LiquidacionAccesoNuevoView(FinancieroMixin, CreateView):
    model = LiquidacionGestor
    form_class = LiquidacionGestorForm
    success_url = "../"
    template_name = "nueva_liquidacion_acceso.html"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        return super(LiquidacionAccesoNuevoView,self).get_context_data(**kwargs)

    def form_valid(self, form):
        self.object = form.save()
        gestor = Gestor.objects.get(pk=self.object.gestor.id)
        gestor.fecha_terminacion = self.object.fecha_terminacion
        gestor.save()
        return super(ModelFormMixin, self).form_valid(form)

class LiquidacionFormacionNuevoView(FinancieroMixin, CreateView):
    model = LiquidacionFormador
    form_class = LiquidacionFormadorForm
    success_url = "../"
    template_name = "nueva_liquidacion_formacion.html"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        return super(LiquidacionFormacionNuevoView,self).get_context_data(**kwargs)

    def form_valid(self, form):
        self.object = form.save()
        formador = Formador.objects.get(pk=self.object.formador.id)
        formador.fecha_terminacion = self.object.fecha_terminacion
        formador.save()
        return super(ModelFormMixin, self).form_valid(form)

class LiquidacionAccesoEditarView(FinancieroMixin, UpdateView):
    model = LiquidacionGestor
    form_class = LiquidacionGestorForm
    success_url = "../../"
    template_name = "nueva_liquidacion_acceso.html"
    pk_url_kwarg = 'id_liquidacion'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        return super(LiquidacionAccesoEditarView,self).get_context_data(**kwargs)

class LiquidacionFormacionEditarView(FinancieroMixin, UpdateView):
    model = LiquidacionFormador
    form_class = LiquidacionFormadorForm
    success_url = "../../"
    template_name = "nueva_liquidacion_formacion.html"
    pk_url_kwarg = 'id_liquidacion'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        return super(LiquidacionFormacionEditarView,self).get_context_data(**kwargs)

def liquidacion_acceso(request,pk,id_liquidacion):
    liquidacion = LiquidacionGestor.objects.get(id=id_liquidacion)

    pythoncom.CoInitializeEx(pythoncom.COINIT_MULTITHREADED)
    xl = win32com.client.dynamic.Dispatch('Word.Application')
    xl.DisplayAlerts = False
    xl.Visible = 0

    archivo = xl.Documents.Open(settings.STATICFILES_DIRS[0]+'/formatos/Formato de liquidaciones.docx')

    archivo.Sections(1).Headers(1).Range.InsertAfter(liquidacion.gestor.nombre.decode('UTF-8').upper()+"\r\r")

    tabla = archivo.Tables(1)
    tabla.Cell(Row=1,Column=2).Range.Text = liquidacion.gestor.nombre.decode('UTF-8').upper()

    if liquidacion.gestor.region.id == 1:
        region = "Región Uno (1) Amazonas, Arauca, Caquetá, Casanare, Guainía, Guaviare, Huila, Meta, Putumayo, Tolima, Vaupés, Vichada".decode('UTF-8')
    if liquidacion.gestor.region.id == 2:
        region = "Región Cuatro (4) Bogotá D.C., Boyacá, Cundinamarca, Norte de Santander, Santander".decode('UTF-8')

    tabla.Cell(Row=2,Column=2).Range.Text = region

    tabla.Cell(Row=3,Column=2).Range.Text = liquidacion.gestor.cedula

    objeto = tabla.Cell(Row=4,Column=2).Range.Text
    objeto = objeto.replace("%REGION%",region)
    tabla.Cell(Row=4,Column=2).Range.Text = objeto

    tabla.Cell(Row=5,Column=2).Range.Text = Money(liquidacion.valor_ejecutado-liquidacion.valor_pagado,'COP').format('es_CO','$#,##0.00')+" Mcte"
    tabla.Cell(Row=7,Column=2).Range.Text = str(int((liquidacion.valor_ejecutado*100.0)/liquidacion.valor_inicial))+" %"
    if liquidacion.gestor.fecha_contratacion != None:
        tabla.Cell(Row=8,Column=2).Range.Text = liquidacion.gestor.fecha_contratacion.strftime("%d/%m/%Y")
    if liquidacion.fecha_terminacion != None:
        tabla.Cell(Row=9,Column=2).Range.Text = liquidacion.fecha_terminacion.strftime("%d/%m/%Y")

    meses = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    mes = int(liquidacion.gestor.fecha_contratacion.strftime("%m"))-1
    fecha = liquidacion.gestor.fecha_contratacion.strftime("%d de "+meses[mes]+" de %Y")

    tabla = archivo.Tables(2)
    tabla.Cell(Row=1,Column=2).Range.Text = fecha

    tabla = archivo.Tables(3)
    tabla.Cell(Row=1,Column=1).Range.Text = liquidacion.gestor.nombre.decode('UTF-8').upper()
    tabla.Cell(Row=1,Column=3).Range.Text = liquidacion.gestor.cedula

    tabla = archivo.Tables(4)
    tabla.Cell(Row=1,Column=2).Range.Text = liquidacion.contrato
    tabla.Cell(Row=1,Column=4).Range.Text = fecha

    tabla = archivo.Tables(5)
    tabla.Cell(Row=1,Column=2).Range.Text = fecha

    tabla = archivo.Tables(6)
    tabla.Cell(Row=1,Column=1).Range.Text = liquidacion.gestor.nombre.decode('UTF-8').upper()
    tabla.Cell(Row=1,Column=3).Range.Text = liquidacion.gestor.cedula

    tabla = archivo.Tables(7)
    tabla.Cell(Row=1,Column=1).Range.Text = "GESTOR TERRITORIAL"

    tabla = archivo.Tables(8)
    tabla.Cell(Row=1,Column=1).Range.Text = "para  Educar  para el periodo 2015 en la "+region

    tabla = archivo.Tables(9)
    tabla.Cell(Row=1,Column=1).Range.Text = fecha

    tabla = archivo.Tables(10)
    tabla.Cell(Row=1,Column=2).Range.Text = Money(liquidacion.valor_inicial,'COP').format('es_CO','$#,##0.00')
    tabla.Cell(Row=2,Column=3).Range.Text = Money(liquidacion.valor_ejecutado,'COP').format('es_CO','$#,##0.00')
    tabla.Cell(Row=3,Column=3).Range.Text = Money(liquidacion.valor_pagado,'COP').format('es_CO','$#,##0.00')

    if liquidacion.valor_ejecutado-liquidacion.valor_pagado >= 0:
        tabla.Cell(Row=4,Column=3).Range.Text = Money(liquidacion.valor_ejecutado-liquidacion.valor_pagado,'COP').format('es_CO','$#,##0.00')
        tabla.Cell(Row=5,Column=3).Range.Text = Money(0,'COP').format('es_CO','$#,##0.00')
    else:
        tabla.Cell(Row=4,Column=3).Range.Text = Money(0,'COP').format('es_CO','$#,##0.00')
        tabla.Cell(Row=5,Column=3).Range.Text = Money(liquidacion.valor_pagado-liquidacion.valor_ejecutado,'COP').format('es_CO','$#,##0.00')

    tabla = archivo.Tables(11)
    tabla.Cell(Row=1,Column=2).Range.Text = Money(liquidacion.valor_ejecutado-liquidacion.valor_pagado,'COP').format('es_CO','$#,##0.00')

    tabla = archivo.Tables(12)
    tabla.Cell(Row=1,Column=2).Range.Text = liquidacion.gestor.nombre.decode('UTF-8').upper()
    tabla.Cell(Row=2,Column=1).Range.Text = liquidacion.gestor.cedula
    tabla.Cell(Row=3,Column=2).Range.Text = liquidacion.contrato
    tabla.Cell(Row=4,Column=2).Range.Text = fecha+"."

    mes = int(datetime.date.today().strftime("%m"))-1
    fecha2 = datetime.date.today().strftime("%d dias del mes de "+meses[mes]+" de %Y")

    tabla = archivo.Tables(13)
    tabla.Cell(Row=1,Column=1).Range.Text = fecha2+"."


    tabla = archivo.Tables(14)
    tabla.Cell(Row=1,Column=2).Range.Text = liquidacion.gestor.nombre.decode('UTF-8').upper()
    tabla.Cell(Row=2,Column=2).Range.Text = "C.C.No "+str(liquidacion.gestor.cedula)


    archivo.SaveAs('C:\\Temp\\Liquidacion '+liquidacion.gestor.nombre.decode('UTF-8')+'.docx')
    archivo.Close(SaveChanges=False)


    zip_subdir = liquidacion.gestor.nombre.decode('UTF-8')
    zip_filename = "%s.zip" % zip_subdir
    s = StringIO.StringIO()
    zf = zipfile.ZipFile(s, "w")

    path = 'C:\\Temp\\Liquidacion '+liquidacion.gestor.nombre.decode('UTF-8')+'.docx'

    if os.path.exists(path):
        fdir, fname = os.path.split(path)
        zf.write(path,os.path.join(liquidacion.gestor.nombre.decode('UTF-8'),fname))

    zf.close()

    os.remove(path)

    resp = HttpResponse(s.getvalue(), content_type = "application/x-zip-compressed")
    resp['Content-Disposition'] = 'attachment; filename=%s' % zip_filename

    return resp

def liquidacion_formacion(request,pk,id_liquidacion):
    liquidacion = LiquidacionFormador.objects.get(id=id_liquidacion)

    pythoncom.CoInitializeEx(pythoncom.COINIT_MULTITHREADED)
    xl = win32com.client.dynamic.Dispatch('Word.Application')
    xl.DisplayAlerts = False
    xl.Visible = 0

    archivo = xl.Documents.Open(settings.STATICFILES_DIRS[0]+'/formatos/Formato de liquidaciones.docx')

    archivo.Sections(1).Headers(1).Range.InsertAfter(liquidacion.formador.nombre.decode('UTF-8').upper()+"\r\r")

    tabla = archivo.Tables(1)
    tabla.Cell(Row=1,Column=2).Range.Text = liquidacion.formador.nombre.decode('UTF-8').upper()

    if liquidacion.formador.region.id == 1:
        region = "Región Uno (1) Amazonas, Arauca, Caquetá, Casanare, Guainía, Guaviare, Huila, Meta, Putumayo, Tolima, Vaupés, Vichada".decode('UTF-8')
    if liquidacion.formador.region.id == 2:
        region = "Región Cuatro (4) Bogotá D.C., Boyacá, Cundinamarca, Norte de Santander, Santander".decode('UTF-8')

    tabla.Cell(Row=2,Column=2).Range.Text = region

    tabla.Cell(Row=3,Column=2).Range.Text = liquidacion.formador.cedula

    objeto = tabla.Cell(Row=4,Column=2).Range.Text
    objeto = objeto.replace("%REGION%",region)
    tabla.Cell(Row=4,Column=2).Range.Text = objeto

    tabla.Cell(Row=5,Column=2).Range.Text = Money(liquidacion.valor_ejecutado-liquidacion.valor_pagado,'COP').format('es_CO','$#,##0.00')+" Mcte"
    tabla.Cell(Row=7,Column=2).Range.Text = str(int((liquidacion.valor_ejecutado*100.0)/liquidacion.valor_inicial))+" %"
    if liquidacion.formador.fecha_contratacion != None:
        tabla.Cell(Row=8,Column=2).Range.Text = liquidacion.formador.fecha_contratacion.strftime("%d/%m/%Y")
    if liquidacion.fecha_terminacion != None:
        tabla.Cell(Row=9,Column=2).Range.Text = liquidacion.fecha_terminacion.strftime("%d/%m/%Y")

    meses = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    mes = int(liquidacion.formador.fecha_contratacion.strftime("%m"))-1
    fecha = liquidacion.formador.fecha_contratacion.strftime("%d de "+meses[mes]+" de %Y")

    tabla = archivo.Tables(2)
    tabla.Cell(Row=1,Column=2).Range.Text = fecha

    tabla = archivo.Tables(3)
    tabla.Cell(Row=1,Column=1).Range.Text = liquidacion.formador.nombre.decode('UTF-8').upper()
    tabla.Cell(Row=1,Column=3).Range.Text = liquidacion.formador.cedula

    tabla = archivo.Tables(4)
    tabla.Cell(Row=1,Column=2).Range.Text = liquidacion.contrato
    tabla.Cell(Row=1,Column=4).Range.Text = fecha

    tabla = archivo.Tables(5)
    tabla.Cell(Row=1,Column=2).Range.Text = fecha

    tabla = archivo.Tables(6)
    tabla.Cell(Row=1,Column=1).Range.Text = liquidacion.formador.nombre.decode('UTF-8').upper()
    tabla.Cell(Row=1,Column=3).Range.Text = liquidacion.formador.cedula

    tabla = archivo.Tables(7)
    tabla.Cell(Row=1,Column=1).Range.Text = "FORMADOR "+liquidacion.formador.tipo.tipo.upper()

    tabla = archivo.Tables(8)
    tabla.Cell(Row=1,Column=1).Range.Text = "para  Educar  para el periodo 2015 en la "+region

    tabla = archivo.Tables(9)
    tabla.Cell(Row=1,Column=1).Range.Text = fecha

    tabla = archivo.Tables(10)
    tabla.Cell(Row=1,Column=2).Range.Text = Money(liquidacion.valor_inicial,'COP').format('es_CO','$#,##0.00')
    tabla.Cell(Row=2,Column=3).Range.Text = Money(liquidacion.valor_ejecutado,'COP').format('es_CO','$#,##0.00')
    tabla.Cell(Row=3,Column=3).Range.Text = Money(liquidacion.valor_pagado,'COP').format('es_CO','$#,##0.00')

    if liquidacion.valor_ejecutado-liquidacion.valor_pagado >= 0:
        tabla.Cell(Row=4,Column=3).Range.Text = Money(liquidacion.valor_ejecutado-liquidacion.valor_pagado,'COP').format('es_CO','$#,##0.00')
        tabla.Cell(Row=5,Column=3).Range.Text = Money(0,'COP').format('es_CO','$#,##0.00')
    else:
        tabla.Cell(Row=4,Column=3).Range.Text = Money(0,'COP').format('es_CO','$#,##0.00')
        tabla.Cell(Row=5,Column=3).Range.Text = Money(liquidacion.valor_pagado-liquidacion.valor_ejecutado,'COP').format('es_CO','$#,##0.00')

    tabla = archivo.Tables(11)
    tabla.Cell(Row=1,Column=2).Range.Text = Money(liquidacion.valor_ejecutado-liquidacion.valor_pagado,'COP').format('es_CO','$#,##0.00')

    tabla = archivo.Tables(12)
    tabla.Cell(Row=1,Column=2).Range.Text = liquidacion.formador.nombre.decode('UTF-8').upper()
    tabla.Cell(Row=2,Column=1).Range.Text = liquidacion.formador.cedula
    tabla.Cell(Row=3,Column=2).Range.Text = liquidacion.contrato
    tabla.Cell(Row=4,Column=2).Range.Text = fecha+"."

    mes = int(datetime.date.today().strftime("%m"))-1
    fecha2 = datetime.date.today().strftime("%d dias del mes de "+meses[mes]+" de %Y")

    tabla = archivo.Tables(13)
    tabla.Cell(Row=1,Column=1).Range.Text = fecha2+"."


    tabla = archivo.Tables(14)
    tabla.Cell(Row=1,Column=2).Range.Text = liquidacion.formador.nombre.decode('UTF-8').upper()
    tabla.Cell(Row=2,Column=2).Range.Text = "C.C.No "+str(liquidacion.formador.cedula)


    archivo.SaveAs('C:\\Temp\\Liquidacion '+liquidacion.formador.nombre.decode('UTF-8')+'.docx')
    archivo.Close(SaveChanges=False)


    zip_subdir = liquidacion.formador.nombre.decode('UTF-8')
    zip_filename = "%s.zip" % zip_subdir
    s = StringIO.StringIO()
    zf = zipfile.ZipFile(s, "w")

    path = 'C:\\Temp\\Liquidacion '+liquidacion.formador.nombre.decode('UTF-8')+'.docx'

    if os.path.exists(path):
        fdir, fname = os.path.split(path)
        zf.write(path,os.path.join(liquidacion.formador.nombre.decode('UTF-8'),fname))

    zf.close()

    os.remove(path)

    resp = HttpResponse(s.getvalue(), content_type = "application/x-zip-compressed")
    resp['Content-Disposition'] = 'attachment; filename=%s' % zip_filename

    return resp