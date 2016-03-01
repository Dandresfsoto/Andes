#!/usr/bin/env python
# -*- coding: utf-8 -*-
from django.shortcuts import render
from django.views.generic import TemplateView, CreateView, UpdateView, FormView
from mixins.mixins import FormacionMixin
from region.models import Region
from formador.models import Formador
from formacion.models import Grupo, ParticipanteEscuelaTic, Entregable, SoporteEntregableEscuelaTic, Masivo, EvidenciaEscuelaTic, Actividad
from formacion.forms import NuevoGrupoForm, NuevoParticipanteForm, NuevoMasivoForm, NuevoSoporteForm, AsignarForm, AgregarSoporteForm
from django.http import HttpResponseRedirect
from django.forms.models import modelformset_factory
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponse
import openpyxl
from django.utils.encoding import smart_unicode
from conf import settings
from django.core.files.base import ContentFile
from StringIO import StringIO
import time
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font
from formacion.models import GrupoDocentes, ParticipanteDocente, SoporteEntregableDocente, ActividadDocentes, EvidenciaDocentes
from formacion.forms import NuevoGrupoDocenteForm, NuevoDocenteForm, AsignarDocenteForm, NuevoSoporteDocenteForm, AgregarSoporteDocenteForm
from pqr.models import PqrRespuesta, Pqr, Llamadas, LlamadasRespuesta
from pqr.forms import PqrRespuestaForm, LlamadaForm, LlamadaRespuestaForm
from django_datatables_view.base_datatable_view import BaseDatatableView
from django.db.models import Q
from django.contrib.auth.models import User
from formacion.models import EvidenciaDocentes, EvidenciaEscuelaTic
import datetime

t = Style(font=Font(name='Calibri',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='C9C9C9',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

co = Style(font=Font(name='Calibri',size=11),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

v = Style(font=Font(name='Calibri',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='E4F5E1',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

vc = Style(font=Font(name='Calibri',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='D4F5CE',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

def validateEmail( email ):
    from django.core.validators import validate_email
    from django.core.exceptions import ValidationError
    try:
        validate_email( email )
        return True
    except ValidationError:
        return False

class FormacionView(FormacionMixin,TemplateView):
    template_name = 'formacion.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        return super(FormacionView,self).get_context_data(**kwargs)

class FormadorView(FormacionMixin,TemplateView):
    template_name = 'tipo2_formador_formacion.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['ID_TIPO'] = 2
        return super(FormadorView,self).get_context_data(**kwargs)

class FormadorTipo1View(FormacionMixin,TemplateView):
    template_name = 'tipo1_formador_formacion.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['ID_TIPO'] = 1
        return super(FormadorTipo1View,self).get_context_data(**kwargs)

class FormadorGrupoView(FormacionMixin,TemplateView):
    template_name = 'tipo2_formador_grupo.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        return super(FormadorGrupoView,self).get_context_data(**kwargs)

class FormadorTipo1GrupoView(FormacionMixin,TemplateView):
    template_name = 'tipo1_formador_grupo.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        return super(FormadorTipo1GrupoView,self).get_context_data(**kwargs)

class NuevoGrupoView(FormacionMixin,CreateView):
    model = Grupo
    form_class = NuevoGrupoForm
    template_name = "formulario_grupo.html"
    success_url = "../"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        return super(NuevoGrupoView,self).get_context_data(**kwargs)

class NuevoGrupoDocenteView(FormacionMixin,CreateView):
    model = GrupoDocentes
    form_class = NuevoGrupoDocenteForm
    template_name = "formulario_grupo_docentes.html"
    success_url = "../"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        return super(NuevoGrupoDocenteView,self).get_context_data(**kwargs)

class FormSoporteGrupoView(FormacionMixin,UpdateView):
    pk_url_kwarg = 'soporte_id'
    model = SoporteEntregableEscuelaTic
    form_class = NuevoSoporteForm
    template_name = "formulario_soporte_tipo2.html"
    success_url = "../../"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        kwargs['NOMBRE_SOPORTE'] = SoporteEntregableEscuelaTic.objects.get(pk=self.kwargs['soporte_id']).entregable.nombre + " - " + SoporteEntregableEscuelaTic.objects.get(pk=self.kwargs['soporte_id']).grupo.nombre
        kwargs['ID_GRUPO'] = self.kwargs['grupo_id']
        kwargs['NOMBRE_GRUPO'] = Grupo.objects.get(pk=self.kwargs['grupo_id']).nombre
        return super(FormSoporteGrupoView,self).get_context_data(**kwargs)

class FormSoporteGrupoDocenteView(FormacionMixin,UpdateView):
    pk_url_kwarg = 'soporte_id'
    model = SoporteEntregableDocente
    form_class = NuevoSoporteDocenteForm
    template_name = "formulario_soporte_tipo1.html"
    success_url = "../../"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        kwargs['NOMBRE_SOPORTE'] = SoporteEntregableDocente.objects.get(pk=self.kwargs['soporte_id']).entregable.nombre + " - " + SoporteEntregableDocente.objects.get(pk=self.kwargs['soporte_id']).grupo.nombre
        kwargs['ID_GRUPO'] = self.kwargs['grupo_id']
        kwargs['NOMBRE_GRUPO'] = GrupoDocentes.objects.get(pk=self.kwargs['grupo_id']).nombre
        return super(FormSoporteGrupoDocenteView,self).get_context_data(**kwargs)

class FormAsignarSoporteView(FormacionMixin,FormView):
    form_class = AsignarForm
    template_name = "formulario_asignar_soporte_tipo2.html"
    success_url = "../../"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        kwargs['NOMBRE_SOPORTE'] = SoporteEntregableEscuelaTic.objects.get(pk=self.kwargs['soporte_id']).entregable.nombre + " - " + SoporteEntregableEscuelaTic.objects.get(pk=self.kwargs['soporte_id']).grupo.nombre
        kwargs['ID_GRUPO'] = self.kwargs['grupo_id']
        kwargs['NOMBRE_GRUPO'] = Grupo.objects.get(pk=self.kwargs['grupo_id']).nombre
        return super(FormAsignarSoporteView,self).get_context_data(**kwargs)

    def get_form_kwargs(self):
        kwargs = {
            'initial': self.get_initial(),
            'prefix': self.get_prefix(),
            'soporte_id': self.kwargs['soporte_id'],
        }
        if self.request.method in ('POST', 'PUT'):
            kwargs.update({
                'data': self.request.POST,
                'files': self.request.FILES,
            })
        return kwargs

    def form_valid(self, form):
        super(FormAsignarSoporteView, self).form_valid(form)
        participantes = form.cleaned_data['participantes']
        soporte = SoporteEntregableEscuelaTic.objects.get(pk=self.kwargs['soporte_id'])

        self.soporte = self.kwargs['soporte_id']
        self.grupo = SoporteEntregableEscuelaTic.objects.get(pk=self.soporte).grupo.id
        self.id_entregable = SoporteEntregableEscuelaTic.objects.get(pk=self.soporte).entregable.id

        borrar = EvidenciaEscuelaTic.objects.filter(soporte__id=self.soporte).exclude(participante__id__in=participantes)
        for evidencia in borrar:
            evidencia.soporte = None
            evidencia.usuario = None
            evidencia.fecha = datetime.datetime.now()
            evidencia.save()

        asignar = EvidenciaEscuelaTic.objects.filter(entregable__id=soporte.entregable.id,participante__id__in=participantes)
        x = asignar.count()
        for evidencia in asignar:
            evidencia.usuario = self.request.user
            evidencia.fecha = datetime.datetime.now()
            evidencia.soporte = soporte
            evidencia.save()

        return HttpResponseRedirect(self.get_success_url())

    def form_invalid(self, form):
        super(FormAsignarSoporteView, self).form_invalid(form)
        soporte = SoporteEntregableEscuelaTic.objects.get(pk=self.kwargs['soporte_id'])

        self.soporte = self.kwargs['soporte_id']
        self.grupo = SoporteEntregableEscuelaTic.objects.get(pk=self.soporte).grupo.id
        self.id_entregable = SoporteEntregableEscuelaTic.objects.get(pk=self.soporte).entregable.id

        participantes_actual = EvidenciaEscuelaTic.objects.filter(soporte__id=self.soporte).values_list("participante__id",flat=True)

        for participante in participantes_actual:
            evidencia = EvidenciaEscuelaTic.objects.filter(participante__id=participante).get(entregable__id=soporte.entregable.id)
            evidencia.soporte = None
            evidencia.save()
        return HttpResponseRedirect(self.get_success_url())

class FormAsignarSoporteDocenteView(FormacionMixin,FormView):
    form_class = AsignarDocenteForm
    template_name = "formulario_asignar_soporte_tipo1.html"
    success_url = "../../"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        kwargs['NOMBRE_SOPORTE'] = SoporteEntregableDocente.objects.get(pk=self.kwargs['soporte_id']).entregable.nombre + " - " + SoporteEntregableDocente.objects.get(pk=self.kwargs['soporte_id']).grupo.nombre
        kwargs['ID_GRUPO'] = self.kwargs['grupo_id']
        kwargs['NOMBRE_GRUPO'] = GrupoDocentes.objects.get(pk=self.kwargs['grupo_id']).nombre
        return super(FormAsignarSoporteDocenteView,self).get_context_data(**kwargs)

    def get_form_kwargs(self):
        kwargs = {
            'initial': self.get_initial(),
            'prefix': self.get_prefix(),
            'soporte_id': self.kwargs['soporte_id'],
        }
        if self.request.method in ('POST', 'PUT'):
            kwargs.update({
                'data': self.request.POST,
                'files': self.request.FILES,
            })
        return kwargs

    def form_valid(self, form):
        super(FormAsignarSoporteDocenteView, self).form_valid(form)
        participantes = form.cleaned_data['participantes']
        soporte = SoporteEntregableDocente.objects.get(pk=self.kwargs['soporte_id'])

        self.soporte = self.kwargs['soporte_id']
        self.grupo = SoporteEntregableDocente.objects.get(pk=self.soporte).grupo.id
        self.id_entregable = SoporteEntregableDocente.objects.get(pk=self.soporte).entregable.id

        #Obtiene los id de los soportes del grupo del entregable especifico y lo convierte en una lista
        x = SoporteEntregableDocente.objects.filter(grupo__formador__id=soporte.grupo.formador.id).filter(entregable__id=self.id_entregable).values_list("id",flat=True)
        x = list(x)

        x.pop(x.index(long(self.soporte)))
        if not isinstance(x,list):
            x = [x]

        # y es una lista con los id de participantes asignados en otro entregable
        y = EvidenciaDocentes.objects.filter(soporte__in=x).values_list("participante__id",flat=True)

        participantes_total = ParticipanteDocente.objects.filter(grupo__formador__id=soporte.grupo.formador.id).values_list("id",flat=True)
        participantes_total = list(set(participantes_total).difference(y))
        participantes_actual = EvidenciaDocentes.objects.filter(soporte__id=self.soporte).values_list("participante__id",flat=True)

        if len(participantes) != 0:
            for participante in participantes_total:
                evidencia = EvidenciaDocentes.objects.filter(participante__id=participante).get(entregable__id=soporte.entregable.id)
                if unicode(participante) in participantes:
                    if evidencia.soporte == None:
                        evidencia.usuario = self.request.user
                        evidencia.fecha = datetime.datetime.now()
                    evidencia.soporte = soporte
                else:
                    evidencia.soporte = None
                    evidencia.usuario = None
                    evidencia.fecha = datetime.datetime.now()
                evidencia.save()
        else:
            for participante in participantes_actual:
                evidencia = EvidenciaDocentes.objects.filter(participante__id=participante).get(entregable__id=soporte.entregable.id)
                if evidencia.soporte == None:
                    evidencia.usuario = self.request.user
                    evidencia.fecha = datetime.datetime.now()
                evidencia.soporte = soporte
                evidencia.usuario = self.request.user
                evidencia.save()
        return HttpResponseRedirect(self.get_success_url())

    def form_invalid(self, form):
        super(FormAsignarSoporteDocenteView, self).form_invalid(form)
        soporte = SoporteEntregableDocente.objects.get(pk=self.kwargs['soporte_id'])

        self.soporte = self.kwargs['soporte_id']
        self.grupo = SoporteEntregableDocente.objects.get(pk=self.soporte).grupo.id
        self.id_entregable = SoporteEntregableDocente.objects.get(pk=self.soporte).entregable.id

        participantes_actual = EvidenciaDocentes.objects.filter(soporte__id=self.soporte).values_list("participante__id",flat=True)

        for participante in participantes_actual:
            evidencia = EvidenciaDocentes.objects.filter(participante__id=participante).get(entregable__id=soporte.entregable.id)
            evidencia.soporte = None
            evidencia.save()
        return HttpResponseRedirect(self.get_success_url())

class FormAgregarSoporteView(FormacionMixin,FormView):
    form_class = AgregarSoporteForm
    template_name = "formulario_agregar_soporte_tipo2.html"
    success_url = "../"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        kwargs['ID_GRUPO'] = self.kwargs['grupo_id']
        kwargs['NOMBRE_GRUPO'] = Grupo.objects.get(pk=self.kwargs['grupo_id']).nombre
        return super(FormAgregarSoporteView,self).get_context_data(**kwargs)

    def form_valid(self, form):
        form.save()
        return HttpResponseRedirect(self.get_success_url())

class FormAgregarSoporteDocenteView(FormacionMixin,FormView):
    form_class = AgregarSoporteDocenteForm
    template_name = "formulario_agregar_soporte_tipo1.html"
    success_url = "../"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        kwargs['ID_GRUPO'] = self.kwargs['grupo_id']
        kwargs['NOMBRE_GRUPO'] = GrupoDocentes.objects.get(pk=self.kwargs['grupo_id']).nombre
        return super(FormAgregarSoporteDocenteView,self).get_context_data(**kwargs)

    def form_valid(self, form):
        form.save()
        return HttpResponseRedirect(self.get_success_url())

class ListadoGrupoView(FormacionMixin,TemplateView):
    template_name = 'tipo2_formador_listado.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        kwargs['ID_GRUPO'] = self.kwargs['grupo_id']
        kwargs['NOMBRE_GRUPO'] = Grupo.objects.get(pk=self.kwargs['grupo_id']).nombre
        return super(ListadoGrupoView,self).get_context_data(**kwargs)

class ListadoGrupoDocentesView(FormacionMixin,TemplateView):
    template_name = 'tipo1_formador_listado.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        kwargs['ID_GRUPO'] = self.kwargs['grupo_id']
        kwargs['NOMBRE_GRUPO'] = GrupoDocentes.objects.get(pk=self.kwargs['grupo_id']).nombre
        return super(ListadoGrupoDocentesView,self).get_context_data(**kwargs)

class NuevoParticipanteView(FormacionMixin,CreateView):
    model = ParticipanteEscuelaTic
    form_class = NuevoParticipanteForm
    template_name = "formulario_participante.html"
    success_url = "../../"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        kwargs['ID_GRUPO'] = self.kwargs['grupo_id']
        kwargs['NOMBRE_GRUPO'] = Grupo.objects.get(pk=self.kwargs['grupo_id']).nombre
        kwargs['ACCION'] = "Nuevo"
        return super(NuevoParticipanteView,self).get_context_data(**kwargs)

class NuevoDocenteView(FormacionMixin,CreateView):
    model = ParticipanteDocente
    form_class = NuevoDocenteForm
    template_name = "formulario_docente.html"
    success_url = "../../"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        kwargs['ID_GRUPO'] = self.kwargs['grupo_id']
        kwargs['NOMBRE_GRUPO'] = GrupoDocentes.objects.get(pk=self.kwargs['grupo_id']).nombre
        kwargs['ACCION'] = "Nuevo"
        kwargs['DEPARTAMENTO_ID'] = GrupoDocentes.objects.get(pk=self.kwargs['grupo_id']).municipio.departamento.id
        return super(NuevoDocenteView,self).get_context_data(**kwargs)

class EditarParticipanteView(FormacionMixin,UpdateView):
    model = ParticipanteEscuelaTic
    form_class = NuevoParticipanteForm
    template_name = "formulario_participante.html"
    success_url = "../../"
    pk_url_kwarg = "participante_id"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        kwargs['ID_GRUPO'] = self.kwargs['grupo_id']
        kwargs['NOMBRE_GRUPO'] = Grupo.objects.get(pk=self.kwargs['grupo_id']).nombre
        kwargs['ACCION'] = "Editar: "+ParticipanteEscuelaTic.objects.get(pk=self.kwargs['participante_id']).nombres
        return super(EditarParticipanteView,self).get_context_data(**kwargs)

class EditarDocenteView(FormacionMixin,UpdateView):
    model = ParticipanteDocente
    form_class = NuevoDocenteForm
    template_name = "formulario_docente.html"
    success_url = "../../"
    pk_url_kwarg = "participante_id"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        kwargs['ID_GRUPO'] = self.kwargs['grupo_id']
        kwargs['NOMBRE_GRUPO'] = GrupoDocentes.objects.get(pk=self.kwargs['grupo_id']).nombre
        kwargs['ACCION'] = "Editar: "+ParticipanteDocente.objects.get(pk=self.kwargs['participante_id']).nombres
        kwargs['DEPARTAMENTO_ID'] = GrupoDocentes.objects.get(pk=self.kwargs['grupo_id']).municipio.departamento.id
        return super(EditarDocenteView,self).get_context_data(**kwargs)

def soporte_form(request,pk,grupo_id,formador_id):
    SoporteFormSet = modelformset_factory(SoporteEntregableEscuelaTic, fields=('soporte',),extra=0)
    if request.method == "POST":
        formset = SoporteFormSet(request.POST, request.FILES, queryset=SoporteEntregableEscuelaTic.objects.filter(grupo__id=grupo_id))
        if formset.is_valid():
            formset.save()
    else:
        formset = SoporteFormSet(queryset=SoporteEntregableEscuelaTic.objects.filter(grupo__id=grupo_id),)
    return render_to_response("soportes_escuelaTIC.html",{"formset":formset,"user":request.user,"REGION":Region.objects.get(pk=pk).nombre},context_instance=RequestContext(request))

class ListadoMasivoView(FormacionMixin,TemplateView):
    template_name = 'tipo2_formador_listado_masivo.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        kwargs['ID_GRUPO'] = self.kwargs['grupo_id']
        kwargs['NOMBRE_GRUPO'] = Grupo.objects.get(pk=self.kwargs['grupo_id']).nombre
        return super(ListadoMasivoView,self).get_context_data(**kwargs)

class NuevoMasivoView(FormacionMixin,CreateView):
    model = Masivo
    form_class = NuevoMasivoForm
    template_name = "formulario_masivo.html"
    success_url = "../"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        kwargs['ID_GRUPO'] = self.kwargs['grupo_id']
        return super(NuevoMasivoView,self).get_context_data(**kwargs)

    def form_valid(self, form):
        self.object = form.save()
        path = smart_unicode(self.object.archivo)
        archivo = openpyxl.load_workbook(settings.MEDIA_ROOT+'/'+path)
        try:
            hoja1 = archivo.get_sheet_by_name('BasePadresFamilia')
        except KeyError:
            self.object.delete()
            return self.render_to_response(self.get_context_data(form=form,ERROR="El Archivo no tiene ninguna hoja con el nombre 'BasePadresFamilia'"))
        else:
            r = StringIO()
            resultado = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

            logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
            logo.drawing.top = 10
            logo.drawing.left = 25

            hoja1 = resultado.get_sheet_by_name('hoja1')
            hoja1.title = "ESCUELATIC"
            hoja1.add_image(logo)

            celda = hoja1.cell('E2')
            celda.value = 'FORMACIÓN'

            celda = hoja1.cell('E3')
            celda.value = 'ESCUELATIC'

            celda = hoja1.cell('I3')
            celda.value = time.strftime("%d/%m/%y")

            celda = hoja1.cell('I4')
            celda.value = time.strftime("%I:%M:%S %p")

            row_num = 5

            columns = [tuple(['NOMBRES',30]),
               tuple(['APELLIDOS',30]),
               tuple(['CEDULA',30]),
               tuple(['RESULTADO',50]),
               ]

            for col_num in xrange(len(columns)):
                c = hoja1.cell(row=row_num, column=col_num+1)
                c.value = columns[col_num][0]
                c.style = t
                hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]

            i=0

            grupo = Grupo.objects.get(pk=self.kwargs['grupo_id'])
            formador = Formador.objects.get(pk=self.kwargs['formador_id'])

            archivo_hoja = archivo.get_sheet_by_name('BasePadresFamilia')
            for fila in archivo_hoja.rows:
                i += 1
                if i > 3:
                    proceso =""
                    try:
                        cedula = fila[6].value.replace('.','').replace(',','')
                    except:
                        cedula = fila[6].value
                    nombres = fila[4].value
                    apellidos = fila[5].value


                    if cedula == "" or cedula == None:
                        proceso = "El campo de cedula esta vacio"
                    else:
                        try:
                            long(cedula)
                        except ValueError:
                            proceso = "El numero de cedula es invalido"
                        else:
                            if ParticipanteEscuelaTic.objects.filter(cedula=cedula).count() != 0:
                                proceso = "La Cedula ya se encuentra registrada"
                            else:
                                if nombres == "" or nombres == None:
                                    proceso = "El campo de nombres esta vacio"
                                else:
                                    if apellidos == "" or apellidos == None:
                                        proceso = "El campo de apellidos esta vacio"
                                    else:
                                        proceso = "Registrado Correctamente"
                                        nuevo = ParticipanteEscuelaTic()
                                        nuevo.grupo = grupo
                                        nuevo.formador = formador
                                        nuevo.numero = fila[0].value
                                        nuevo.departamento = grupo.municipio.departamento.nombre
                                        nuevo.municipio = grupo.municipio.nombre
                                        nuevo.institucion = fila[3].value
                                        nuevo.nombres = fila[4].value
                                        nuevo.apellidos = fila[5].value
                                        nuevo.cedula = long(cedula)
                                        if fila[7].value == None or fila[7].value == "":
                                            nuevo.genero = "Masculino"
                                            proceso = "Registrado Correctamente - Genero Masculino por defecto"
                                        else:
                                            nuevo.genero = fila[7].value
                                        nuevo.nivel_educativo = fila[8].value
                                        nuevo.telefono = fila[9].value
                                        if validateEmail(fila[10].value):
                                            nuevo.correo = fila[10].value
                                        nuevo.poblacion = fila[11].value
                                        nuevo.codigo_anspe = fila[12].value
                                        nuevo.tipo_proyecto = fila[13].value
                                        nuevo.grupo_conformacion = fila[14].value
                                        nuevo.save()


                    row_num += 1
                    row = [
                        fila[4].value,
                        fila[5].value,
                        fila[6].value,
                        proceso
                    ]

                    for col_num in xrange(len(row)):
                        c = hoja1.cell(row=row_num, column=col_num+1)
                        if row[col_num] == True:
                            c.value = "SI"
                        if row[col_num] == False:
                            c.value = "NO"
                        if row[col_num] == None:
                            c.value = ""
                        else:
                            c.value = row[col_num]
                        c.style = co

            resultado.save(r)
            self.object.resultado.save('Resultado.xlsx', ContentFile(r.getvalue()))
            self.object.save()
        return HttpResponseRedirect(self.get_success_url())

class CalificarGrupoView(FormacionMixin,TemplateView):
    template_name = 'tipo2_formador_grupo_calificar.html'

    def get_context_data(self, **kwargs):
        participantes = ParticipanteEscuelaTic.objects.filter(grupo__id=self.kwargs['grupo_id']).count()
        soportes = SoporteEntregableEscuelaTic.objects.filter(grupo__id=self.kwargs['grupo_id']).order_by('entregable__id')
        id_actividades = soportes.filter(entregable__actividad__id=1).values_list('entregable__actividad__id',flat=True)
        id_actividades = list(set(id_actividades))
        y=[]
        i=0
        for id_actividad in id_actividades:
            x=[]
            soportes_filtro = soportes.filter(entregable__actividad__id=id_actividad)
            nombre_actividad = Actividad.objects.get(id=id_actividad).nombre
            maximo = max(soportes_filtro.values_list('id',flat=True))
            for soporte_filtro in soportes_filtro:
                i += 1
                if i%2 == 0:
                    clase = "even"
                else:
                    clase = "odd"
                if soporte_filtro.id == maximo:
                    clase ="max"
                cantidad = EvidenciaEscuelaTic.objects.filter(soporte_id=soporte_filtro.id).count()
                x.append({"actividad":soporte_filtro.entregable.actividad.nombre,"entregable":soporte_filtro.entregable.nombre ,"id_soporte" : soporte_filtro.id ,"link_soporte" : str(soporte_filtro.soporte),"cantidad":cantidad,"clase":clase})
            y.append({"nombre_actividad":nombre_actividad,"informacion":x})

        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        kwargs['ID_GRUPO'] = self.kwargs['grupo_id']
        kwargs['NOMBRE_GRUPO'] = Grupo.objects.get(pk=self.kwargs['grupo_id']).nombre
        kwargs['ENTREGABLES'] = y
        return super(CalificarGrupoView,self).get_context_data(**kwargs)

class CalificarGrupoDocentesView(FormacionMixin,TemplateView):
    template_name = 'tipo1_formador_grupo_calificar.html'

    def get_context_data(self, **kwargs):
        participantes = ParticipanteDocente.objects.filter(grupo__id=self.kwargs['grupo_id']).count()
        #excluidos = [9,11,12,13,15,17,18,19,20,25,26,27,35,36,37,38,39,40,51,53,55,59]
        excluidos = []
        soportes = SoporteEntregableDocente.objects.filter(grupo__id=self.kwargs['grupo_id']).order_by('entregable__id')
        id_actividades = soportes.values_list('entregable__actividad__id',flat=True)
        id_actividades = list(set(id_actividades))
        y=[]
        i=0
        for id_actividad in id_actividades:
            x=[]
            soportes_filtro = soportes.filter(entregable__actividad__id=id_actividad)
            nombre_actividad = ActividadDocentes.objects.get(id=id_actividad).nombre
            maximo = max(soportes_filtro.values_list('id',flat=True))
            for soporte_filtro in soportes_filtro:
                i += 1
                if i%2 == 0:
                    clase = "even"
                else:
                    clase = "odd"
                if soporte_filtro.id == maximo:
                    clase ="max"
                cantidad = EvidenciaDocentes.objects.filter(soporte_id=soporte_filtro.id).count()
                x.append({"actividad":soporte_filtro.entregable.actividad.nombre,"entregable":soporte_filtro.entregable.nombre ,"id_soporte" : soporte_filtro.id ,"link_soporte" : str(soporte_filtro.soporte),"cantidad":cantidad,"clase":clase})
            y.append({"nombre_actividad":nombre_actividad,"informacion":x})

        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['NOMBRE_FORMADOR'] = Formador.objects.get(pk=self.kwargs['formador_id']).nombre
        kwargs['ID_FORMADOR'] = self.kwargs['formador_id']
        kwargs['ID_GRUPO'] = self.kwargs['grupo_id']
        kwargs['NOMBRE_GRUPO'] = GrupoDocentes.objects.get(pk=self.kwargs['grupo_id']).nombre
        kwargs['ENTREGABLES'] = y
        return super(CalificarGrupoDocentesView,self).get_context_data(**kwargs)

class MapView(FormacionMixin,TemplateView):
    template_name = 'map.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = Region.objects.get(pk=self.kwargs['pk']).id
        return super(MapView,self).get_context_data(**kwargs)

class MapRespuestaView(FormacionMixin,TemplateView):
    template_name = 'map_respuesta.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = Region.objects.get(pk=self.kwargs['pk']).id
        kwargs['ID_CODIGO'] = Pqr.objects.get(pk=self.kwargs['codigo']).id
        return super(MapRespuestaView,self).get_context_data(**kwargs)

class MapNuevaRespuestaView(FormacionMixin,CreateView):
    model = PqrRespuesta
    form_class = PqrRespuestaForm
    template_name = 'map_nueva_respuesta.html'
    success_url = "../"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = Region.objects.get(pk=self.kwargs['pk']).id
        kwargs['ID_CODIGO'] = Pqr.objects.get(pk=self.kwargs['codigo']).id
        return super(MapNuevaRespuestaView,self).get_context_data(**kwargs)

class MapEditarRespuestaView(FormacionMixin,UpdateView):
    model = PqrRespuesta
    form_class = PqrRespuestaForm
    template_name = 'map_editar_respuesta.html'
    success_url = "../../"
    pk_url_kwarg = "codigo_respuesta"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = Region.objects.get(pk=self.kwargs['pk']).id
        kwargs['ID_CODIGO'] = Pqr.objects.get(pk=self.kwargs['codigo']).id
        return super(MapEditarRespuestaView,self).get_context_data(**kwargs)

class LlamadasView(FormacionMixin,TemplateView):
    template_name = 'llamadas.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = Region.objects.get(pk=self.kwargs['pk']).id
        return super(LlamadasView,self).get_context_data(**kwargs)

class LlamadasRespuestaView(FormacionMixin,TemplateView):
    template_name = 'llamadas_respuesta.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = Region.objects.get(pk=self.kwargs['pk']).id
        kwargs['ID_CODIGO'] = Llamadas.objects.get(pk=self.kwargs['codigo']).id
        return super(LlamadasRespuestaView,self).get_context_data(**kwargs)

class LlamadasNuevaRespuestaView(FormacionMixin,CreateView):
    model = LlamadasRespuesta
    form_class = LlamadaRespuestaForm
    template_name = 'llamadas_nueva_respuesta.html'
    success_url = "../"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = Region.objects.get(pk=self.kwargs['pk']).id
        kwargs['ID_CODIGO'] = Llamadas.objects.get(pk=self.kwargs['codigo']).id
        return super(LlamadasNuevaRespuestaView,self).get_context_data(**kwargs)

class LlamadasNuevoView(FormacionMixin,CreateView):
    model = Llamadas
    form_class = LlamadaForm
    template_name = 'llamadas_nueva.html'
    success_url = "../"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = Region.objects.get(pk=self.kwargs['pk']).id
        return super(LlamadasNuevoView,self).get_context_data(**kwargs)

class ListaAuxiliaresView(BaseDatatableView):
    model = User
    columns = [
        'id',
        'username',
        'first_name',
        'last_name'
    ]

    order_columns = [
        'id',
        'id',
        'id',
        'id'
    ]

    def get_initial_queryset(self):
        if not self.model:
            raise NotImplementedError("Need to provide a model or implement get_initial_queryset!")
        y = EvidenciaDocentes.objects.exclude(usuario=None).values_list('usuario__id',flat=True)
        z = EvidenciaEscuelaTic.objects.exclude(usuario=None).values_list('usuario__id',flat=True)
        return self.model.objects.filter(id__in=list(y)+list(z))

    def filter_queryset(self, qs):
        search = self.request.GET.get(u'search[value]', None)
        q = Q()
        if search:
            qs = qs.filter(q)
        return qs

    def prepare_results(self, qs):
        json_data = []
        for item in qs:
            hoy = datetime.datetime.now().date()
            manana = hoy + datetime.timedelta(1)
            hoy_start = datetime.datetime.combine(hoy, datetime.time())
            hoy_end = datetime.datetime.combine(manana, datetime.time())
            cantidad_padres_dia = EvidenciaEscuelaTic.objects.filter(usuario=item.id).filter(fecha__lte=hoy_end,fecha__gte=hoy_start).count()
            cantidad_docentes_dia = EvidenciaDocentes.objects.filter(usuario=item.id).filter(fecha__lte=hoy_end,fecha__gte=hoy_start).count()
            cantidad_docentes = EvidenciaDocentes.objects.filter(usuario=item.id).count()
            cantidad_padres = EvidenciaEscuelaTic.objects.filter(usuario=item.id).count()
            json_data.append([
                item.id,
                item.username,
                item.first_name,
                item.last_name,
                cantidad_docentes,
                cantidad_padres,
                cantidad_docentes_dia,
                cantidad_padres_dia
            ])

        return json_data