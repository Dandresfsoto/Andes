#!/usr/bin/env python
# -*- coding: utf-8 -*-
from django.views.generic import ListView
from django.views.generic import TemplateView
from .models import Region
from eje.models import Eje
from formador.models import Formador
from mixins.mixins import RegionMixin, AndesMixin, CpeMixin
from django.shortcuts import HttpResponseRedirect
from django.http import HttpResponse
import os
import zipfile
import StringIO
from django.conf import settings

from funcionario.models import Funcionario
from gestor.models import Gestor
import datetime
from string import maketrans
import time
import datetime
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font
import openpyxl
from acceso.models import Evidencia, EvidenciaApoyo
from radicado.models import Radicado

t = Style(font=Font(name='Calibri',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='C9C9C9',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

co = Style(font=Font(name='Calibri',size=11),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

v = Style(font=Font(name='Calibri',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='E4F5E1',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

vc = Style(font=Font(name='Calibri',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='D4F5CE',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

def encode_cp437(s, _noqmarks=maketrans('?', '_')):
    return s.encode('cp437', errors='replace').translate(_noqmarks)

class InicioView(ListView):
    template_name = 'inicio.html'
    model = Region

    def get_context_data(self, **kwargs):
        return super(InicioView,self).get_context_data(**kwargs)

class RegionView(RegionMixin, ListView):
    template_name = 'region.html'
    model = Region

    def queryset(self):
        queryset = Region.objects.get(pk=self.kwargs['pk'])
        return queryset

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        return super(RegionView,self).get_context_data(**kwargs)

class AndesView(AndesMixin, ListView):
    template_name = 'andes.html'
    model = Region

    def queryset(self):
        queryset = Region.objects.get(pk=self.kwargs['pk'])
        return queryset

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        return super(AndesView,self).get_context_data(**kwargs)

class CpeView(CpeMixin,ListView):
    template_name = 'cpe.html'
    model = Region

    def queryset(self):
        queryset = Region.objects.get(pk=self.kwargs['pk'])
        return queryset

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        return super(CpeView,self).get_context_data(**kwargs)


class CpeFormacionView(CpeMixin,ListView):
    template_name = 'formacion_cpe.html'
    model = Region

    def queryset(self):
        queryset = Region.objects.get(pk=self.kwargs['pk'])
        return queryset

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        return super(CpeFormacionView,self).get_context_data(**kwargs)

class CpeAccesoView(CpeMixin,ListView):
    template_name = 'acceso_cpe.html'
    model = Region

    def queryset(self):
        queryset = Region.objects.get(pk=self.kwargs['pk'])
        return queryset

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        return super(CpeAccesoView,self).get_context_data(**kwargs)

class CpeFuncionarioView(CpeMixin,TemplateView):
    template_name = 'listado_funcionarios_cpe.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['EJE'] = Eje.objects.all().filter(nombre=self.kwargs['eje'])[0].nombre
        kwargs['ID_EJE'] = Eje.objects.all().filter(nombre=self.kwargs['eje'])[0].id
        return super(CpeFuncionarioView,self).get_context_data(**kwargs)

class CpeGestorView(CpeMixin,TemplateView):
    template_name = 'listado_gestores_cpe.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['ID_TIPO'] = 1
        return super(CpeGestorView,self).get_context_data(**kwargs)

class CpeGestorApoyoView(CpeMixin,TemplateView):
    template_name = 'listado_gestores_cpe.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['ID_TIPO'] = 2
        return super(CpeGestorApoyoView,self).get_context_data(**kwargs)

class CpeFormadorView(CpeMixin,TemplateView):
    template_name = 'tipo_formador_cpe.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        return super(CpeFormadorView,self).get_context_data(**kwargs)

class CpeFormadorTipoView(CpeMixin,TemplateView):
    template_name = 'listado_formadores_cpe.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['ID_TIPO'] = self.kwargs['tipo']
        return super(CpeFormadorTipoView,self).get_context_data(**kwargs)

class CpeAdministrativoView(CpeMixin,TemplateView):
    template_name = 'administrativo_cpe.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        return super(CpeAdministrativoView,self).get_context_data(**kwargs)

class CpeAdministrativoInformesView(CpeMixin,TemplateView):
    template_name = 'administrativo_cpe_listado.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        return super(CpeAdministrativoInformesView,self).get_context_data(**kwargs)

class CpeAdministrativoObligacionesView(CpeMixin,TemplateView):
    template_name = 'administrativo_cpe_listado_obligaciones.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        return super(CpeAdministrativoObligacionesView,self).get_context_data(**kwargs)

def hv(request,pk,tipo):
    formadores = Formador.objects.filter(region__id=pk).filter(tipo__id=tipo)
    zip_subdir = "Hojas de Vida"
    zip_filename = "%s.zip" % zip_subdir
    zf = zipfile.ZipFile(settings.MEDIA_ROOT+'\\Formadores\\Hojas de Vida\\'+tipo+'\\'+zip_filename, mode='w',allowZip64 = True)

    for formador in formadores:
        soporte = settings.MEDIA_ROOT+'/'+str(formador.hv)
        if os.path.exists(soporte):
            fdir, fname = os.path.split(soporte)
            zf.write(soporte,os.path.join('Hoja de Vida',encode_cp437(formador.nombre),encode_cp437(fname)))

    zf.close()
    return HttpResponseRedirect('/media/Formadores/Hojas de Vida/'+tipo+'/'+zip_filename)

def contratos(request,pk,tipo):
    formadores = Formador.objects.filter(region__id=pk).filter(tipo__id=tipo)
    zip_subdir = "Contratos"
    zip_filename = "%s.zip" % zip_subdir
    zf = zipfile.ZipFile(settings.MEDIA_ROOT+'\\Formadores\\Contratos\\'+tipo+'\\'+zip_filename, mode='w',allowZip64 = True)

    for formador in formadores:
        soporte = settings.MEDIA_ROOT+'/'+str(formador.contrato)
        if os.path.exists(soporte):
            fdir, fname = os.path.split(soporte)
            zf.write(soporte,os.path.join('Contratos',encode_cp437(formador.nombre),encode_cp437(fname)))

    zf.close()
    return HttpResponseRedirect('/media/Formadores/Contratos/'+tipo+'/'+zip_filename)

def hvFuncionarios(request,pk,eje):
    funcionarios = Funcionario.objects.filter(region__id=pk).filter(eje__nombre=eje)
    zip_subdir = "Hojas de Vida"
    zip_filename = "%s.zip" % zip_subdir
    zf = zipfile.ZipFile(settings.MEDIA_ROOT+'\\Funcionarios\\Hojas de Vida\\'+zip_filename, mode='w',allowZip64 = True)

    for funcionario in funcionarios:
        soporte = settings.MEDIA_ROOT+'/'+str(funcionario.hv)
        if os.path.exists(soporte):
            fdir, fname = os.path.split(soporte)
            zf.write(soporte,os.path.join('Hoja de Vida',encode_cp437(funcionario.nombre),encode_cp437(fname)))

    zf.close()
    return HttpResponseRedirect('/media/Funcionarios/Hojas de Vida/'+zip_filename)

def contratosFuncionarios(request,pk,eje):
    funcionarios = Funcionario.objects.filter(region__id=pk).filter(eje__nombre=eje)
    zip_subdir = "Contratos"
    zip_filename = "%s.zip" % zip_subdir
    zf = zipfile.ZipFile(settings.MEDIA_ROOT+'\\Funcionarios\\Contratos\\'+zip_filename, mode='w',allowZip64 = True)

    for funcionario in funcionarios:
        soporte = settings.MEDIA_ROOT+'/'+str(funcionario.contrato)
        if os.path.exists(soporte):
            fdir, fname = os.path.split(soporte)
            zf.write(soporte,os.path.join('Contratos',encode_cp437(funcionario.nombre),encode_cp437(fname)))

    zf.close()
    return HttpResponseRedirect('/media/Funcionarios/Contratos/'+zip_filename)

def hvGestores(request,pk,tipo):
    gestores = Gestor.objects.filter(region__id=pk).filter(tipo__id=tipo)
    zip_subdir = "Hojas de Vida"
    zip_filename = "%s.zip" % zip_subdir
    zf = zipfile.ZipFile(settings.MEDIA_ROOT+'\\HV\\'+tipo+'\\'+zip_filename, mode='w',allowZip64 = True)

    for gestor in gestores:
        soporte = settings.MEDIA_ROOT+'/'+str(gestor.hv)
        if os.path.exists(soporte):
            fdir, fname = os.path.split(soporte)
            zf.write(soporte,os.path.join('Hoja de Vida',encode_cp437(gestor.nombre),encode_cp437(fname)))

    zf.close()
    return HttpResponseRedirect('/media/HV/'+tipo+'/'+zip_filename)

def contratosGestores(request,pk,tipo):
    gestores = Gestor.objects.filter(region__id=pk).filter(tipo__id=tipo)
    zip_subdir = "Contratos"
    zip_filename = "%s.zip" % zip_subdir
    zf = zipfile.ZipFile(settings.MEDIA_ROOT+'\\Contratos\\'+tipo+'\\'+zip_filename, mode='w',allowZip64 = True)

    for gestor in gestores:
        soporte = settings.MEDIA_ROOT+'/'+str(gestor.contrato)
        if os.path.exists(soporte):
            fdir, fname = os.path.split(soporte)
            zf.write(soporte,os.path.join('Contratos',encode_cp437(gestor.nombre),encode_cp437(fname)))
    zf.close()
    return HttpResponseRedirect('/media/Contratos/'+tipo+'/'+zip_filename)

def ruteoGestores(request,pk,tipo):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Ruteo Gestores.xlsx'
    archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

    logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
    logo.drawing.top = 10
    logo.drawing.left = 25

    hoja1 = archivo.get_sheet_by_name('hoja1')
    hoja1.title = "Ruteo Gestores Territoriales"
    hoja1.add_image(logo)

    celda = hoja1.cell('E2')
    celda.value = 'ACCESO'

    celda = hoja1.cell('E3')
    celda.value = 'RUTEO GESTORES TERRITORIALES'

    celda = hoja1.cell('I3')
    celda.value = time.strftime("%d/%m/%y")

    celda = hoja1.cell('I4')
    celda.value = time.strftime("%I:%M:%S %p")

    row_num = 5

    columns = [tuple(['Region',30]),
               tuple(['Radicado',30]),
               tuple(['Departamento',30]),
               tuple(['Municipio',30]),
               tuple(['Nombre Institución',30]),
               tuple(['Dane Institución',30]),
               tuple(['Nombre Sede',30]),
               tuple(['Dane Sede',30]),
               tuple(['Ubicación',30]),
               tuple(['Nombre Gestor',30]),
               tuple(['Cedula',30]),
               tuple(['Celular',30]),
               tuple(['Correo',30]),
               ]



    for col_num in xrange(len(columns)):
        c = hoja1.cell(row=row_num, column=col_num+1)
        c.value = columns[col_num][0]
        c.style = t
        hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


    gestores = Gestor.objects.filter(region__id=pk).filter(tipo__id=tipo)

    for gestor in gestores:
        radicados = Evidencia.objects.filter(gestor__id=gestor.id).values_list('radicado__id',flat=True).distinct()
        for radicado in radicados:
            r = Radicado.objects.get(pk=radicado)
            row_num += 1
            row = [
                gestor.region.nombre,
                r.numero,
                r.municipio.departamento.nombre,
                r.municipio.nombre,
                r.nombre_institucion,
                r.dane_institucion,
                r.nombre_sede,
                r.dane_sede,
                r.zona,
                gestor.nombre,
                gestor.cedula,
                gestor.celular,
                gestor.correo
            ]

            for col_num in xrange(len(row)):
                c = hoja1.cell(row=row_num, column=col_num+1)
                if row[col_num] == True:
                    c.value = "SI"
                if row[col_num] == False:
                    c.value = "NO"
                if row[col_num] == None:
                    c.value = ""
                else:
                    c.value = row[col_num]
                c.style = co



    archivo.save(response)
    return response

def ruteoGestoresApoyo(request,pk,tipo):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Ruteo Gestores.xlsx'
    archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

    logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
    logo.drawing.top = 10
    logo.drawing.left = 25

    hoja1 = archivo.get_sheet_by_name('hoja1')
    hoja1.title = "Ruteo Gestores Apoyo"
    hoja1.add_image(logo)

    celda = hoja1.cell('E2')
    celda.value = 'ACCESO'

    celda = hoja1.cell('E3')
    celda.value = 'RUTEO GESTORES APOYO'

    celda = hoja1.cell('I3')
    celda.value = time.strftime("%d/%m/%y")

    celda = hoja1.cell('I4')
    celda.value = time.strftime("%I:%M:%S %p")

    row_num = 5

    columns = [tuple(['Region',30]),
               tuple(['Radicado',30]),
               tuple(['Departamento',30]),
               tuple(['Municipio',30]),
               tuple(['Nombre Institución',30]),
               tuple(['Dane Institución',30]),
               tuple(['Nombre Sede',30]),
               tuple(['Dane Sede',30]),
               tuple(['Ubicación',30]),
               tuple(['Nombre Gestor',30]),
               tuple(['Cedula',30]),
               tuple(['Celular',30]),
               tuple(['Correo',30]),
               ]



    for col_num in xrange(len(columns)):
        c = hoja1.cell(row=row_num, column=col_num+1)
        c.value = columns[col_num][0]
        c.style = t
        hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


    gestores = Gestor.objects.filter(region__id=pk).filter(tipo__id=tipo)

    for gestor in gestores:
        radicados = EvidenciaApoyo.objects.filter(gestor__id=gestor.id).values_list('radicado__id',flat=True).distinct()
        for radicado in radicados:
            r = Radicado.objects.get(pk=radicado)
            row_num += 1
            row = [
                gestor.region.nombre,
                r.numero,
                r.municipio.departamento.nombre,
                r.municipio.nombre,
                r.nombre_institucion,
                r.dane_institucion,
                r.nombre_sede,
                r.dane_sede,
                r.zona,
                gestor.nombre,
                gestor.cedula,
                gestor.celular,
                gestor.correo
            ]

            for col_num in xrange(len(row)):
                c = hoja1.cell(row=row_num, column=col_num+1)
                if row[col_num] == True:
                    c.value = "SI"
                if row[col_num] == False:
                    c.value = "NO"
                if row[col_num] == None:
                    c.value = ""
                else:
                    c.value = row[col_num]
                c.style = co



    archivo.save(response)
    return response